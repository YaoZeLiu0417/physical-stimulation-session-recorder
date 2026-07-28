import ast
from collections.abc import Mapping, Sequence
from copy import deepcopy
from dataclasses import FrozenInstanceError, replace
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
import json
from pathlib import Path
import traceback
from zipfile import ZipFile

import openpyxl
import pytest

from app_workflow import daily_context_values
import questionnaire_export
from questionnaire_export import (
    ParticipantSnapshot,
    build_participant_export,
    build_participant_snapshot,
    participant_snapshot_json,
    questionnaire_export_sheets,
)
from questionnaire_specs import (
    DAILY_CONDITIONAL,
    DAILY_CORE,
    FORMAL_INSTRUMENTS,
    VISIT_INSTRUMENT_IDS,
    WEEKLY_INSTRUMENTS,
)
from session_record_workflow import (
    create_session_record,
    mark_questionnaire_visit_complete,
    persist_daily_questionnaire,
    persist_formal_questionnaire,
)


MODULE_SOURCE = Path(__file__).resolve().parents[1] / "questionnaire_export.py"
EXPORTED_AT = datetime(2026, 7, 27, 10, 30, tzinfo=timezone.utc)
EXPORTED_AT_ISO = "2026-07-27T10:30:00+00:00"
CREATED_AT_ISO = "2026-07-27T10:00:00+00:00"
COMPLETED_AT_ISO = "2026-07-27T10:15:00Z"
RECORDING_KEYS = (
    "version",
    "storage",
    "status",
    "mode",
    "duration_seconds",
    "camera_ready",
    "microphone_ready",
    "saved_confirmed",
)
INSTRUMENT_VERSION_KEYS = (
    "daily_nssi_ema",
    "weekly_nssi",
    "formal_nssi_crf",
)


def _raw_value(question, *, positive: bool) -> object:
    if question.kind == "boolean":
        return positive
    if question.kind in {"slider", "integer"}:
        return question.min_value
    if question.kind == "multiselect":
        return list(question.options[:2])
    if question.kind == "text":
        return f"raw:{question.id}"
    raise AssertionError(f"unsupported question kind: {question.kind}")


def _recording(status: str = "saved") -> dict[str, object]:
    saved = status == "saved"
    return {
        "version": 2,
        "storage": "browser_local",
        "status": status,
        "mode": "long" if saved else "demo",
        "duration_seconds": 1250 if saved else 0,
        "camera_ready": False,
        "microphone_ready": False,
        "saved_confirmed": saved,
    }


def _daily_questions(day: int):
    weekly = (
        tuple(
            question
            for instrument in WEEKLY_INSTRUMENTS
            for question in instrument.questions
        )
        if day in {7, 14, 21, 28}
        else ()
    )
    return (*DAILY_CORE, *DAILY_CONDITIONAL, *weekly)


def _formal_questions(visit: str):
    return tuple(
        question
        for instrument_id in VISIT_INSTRUMENT_IDS[visit]
        for question in FORMAL_INSTRUMENTS[instrument_id].questions
    )


def _completed_record(
    *,
    day: int = 1,
    visit: str = "daily",
    positive: bool = False,
    recording_status: str = "saved",
) -> dict[str, object]:
    record = create_session_record(
        "sub-001",
        date(2026, 7, 27),
        day,
        visit,
        token="01abcdef",
        now_iso=CREATED_AT_ISO,
    )
    if visit == "daily":
        questions = _daily_questions(day)
        answers = {
            question.id: _raw_value(question, positive=positive)
            for question in questions
        }
        persist_daily_questionnaire(
            record,
            answers,
            set(answers),
            current_step=max(0, len(answers) - 1),
            daily_context={
                "sleep_hours": 7.5,
                "mood_1to9": 6,
                "stress_1to9": 3,
                "pain_0to10": 0,
                "nssi_urge_0to10": 2,
                "coping_effect_1to5": 4,
                "caffeine": "少量",
                "exercise": "中等",
                "tags": ["home", "evening"],
                "coping_used": ["breathing"],
                "narrative": "raw narrative",
                "triggers": "raw trigger",
            },
        )
    else:
        questions = _formal_questions(visit)
        answers = {
            question.id: _raw_value(question, positive=True)
            for question in questions
        }
        persist_formal_questionnaire(
            record,
            visit,
            answers,
            set(answers),
            current_step=max(0, len(answers) - 1),
        )
    record["recording"] = _recording(recording_status)
    mark_questionnaire_visit_complete(
        record,
        visit,
        completed_at_iso=COMPLETED_AT_ISO,
    )
    return record


def _bundle_parts(record: dict[str, object], *, visit: str):
    bundle = build_participant_export(
        record,
        visit=visit,
        exported_at=EXPORTED_AT,
    )
    with ZipFile(BytesIO(bundle.data), "r") as archive:
        names = archive.namelist()
        parsed_json = json.loads(archive.read("responses.json"))
        workbook_bytes = archive.read("responses.xlsx")
    workbook = openpyxl.load_workbook(BytesIO(workbook_bytes), data_only=False)
    return bundle, names, parsed_json, workbook


def _worksheet_rows(worksheet) -> list[dict[str, object]]:
    rows = list(worksheet.iter_rows(values_only=True))
    headers = rows[0]
    return [dict(zip(headers, row, strict=True)) for row in rows[1:]]


def _assert_no_sentinel_in_exception(error: BaseException, sentinel: str) -> None:
    rendered = "".join(
        traceback.TracebackException.from_exception(error).format(chain=True)
    )
    assert sentinel not in rendered
    seen: set[int] = set()
    current: BaseException | None = error
    while current is not None and id(current) not in seen:
        seen.add(id(current))
        assert sentinel not in str(current)
        current = current.__cause__ or current.__context__


class _ExplodingMapping(Mapping[str, object]):
    def __init__(
        self,
        source: Mapping[str, object],
        *,
        sentinel: str,
        operation: str,
        key: str = "",
    ) -> None:
        self._source = source
        self._sentinel = sentinel
        self._operation = operation
        self._key = key

    def __getitem__(self, key: str) -> object:
        if self._operation == "getitem" and key == self._key:
            raise RuntimeError(self._sentinel)
        return self._source[key]

    def __iter__(self):
        if self._operation == "iter":
            raise RuntimeError(self._sentinel)
        return iter(self._source)

    def __len__(self) -> int:
        if self._operation == "len":
            raise RuntimeError(self._sentinel)
        return len(self._source)


class _ExplodingSequence(Sequence[str]):
    def __init__(
        self,
        source: Sequence[str],
        *,
        sentinel: str,
        operation: str,
    ) -> None:
        self._source = source
        self._sentinel = sentinel
        self._operation = operation

    def __getitem__(self, index):
        if self._operation == "getitem":
            raise RuntimeError(self._sentinel)
        return self._source[index]

    def __iter__(self):
        if self._operation == "iter":
            raise RuntimeError(self._sentinel)
        return iter(self._source)

    def __len__(self) -> int:
        if self._operation == "len":
            raise RuntimeError(self._sentinel)
        return len(self._source)


class _SwitchingMapping(Mapping[str, object]):
    def __init__(
        self,
        source: Mapping[str, object],
        *,
        later_values: Mapping[str, object],
    ) -> None:
        self._source = source
        self._later_values = later_values
        self.read_counts: dict[str, int] = {}

    def __getitem__(self, key: str) -> object:
        self.read_counts[key] = self.read_counts.get(key, 0) + 1
        if self.read_counts[key] > 1 and key in self._later_values:
            return self._later_values[key]
        return self._source[key]

    def __iter__(self):
        return iter(self._source)

    def __len__(self) -> int:
        return len(self._source)


class _EqualitySpoof(str):
    def __new__(cls, visible_value: str, accepted_value: str):
        instance = super().__new__(cls, visible_value)
        instance.accepted_value = accepted_value
        return instance

    def __eq__(self, other: object) -> bool:
        return other == self.accepted_value

    def __ne__(self, other: object) -> bool:
        return not self == other

    def __hash__(self) -> int:
        return hash(self.accepted_value)


class _EqualitySpoofInt(int):
    def __new__(cls, visible_value: int, accepted_value: int):
        instance = super().__new__(cls, visible_value)
        instance.accepted_value = accepted_value
        return instance

    def __eq__(self, other: object) -> bool:
        return other == self.accepted_value

    def __ne__(self, other: object) -> bool:
        return not self == other

    def __hash__(self) -> int:
        return hash(self.accepted_value)


def test_day_one_negative_branch_builds_canonical_frozen_snapshot() -> None:
    record = _completed_record()

    snapshot = build_participant_snapshot(
        record,
        visit="daily",
        exported_at_iso=EXPORTED_AT_ISO,
    )

    assert snapshot.export_schema_version == 1
    assert snapshot.participant_id == "sub-001"
    assert snapshot.record_date == "2026-07-27"
    assert snapshot.intervention_day == 1
    assert snapshot.visit == "daily"
    assert snapshot.exported_at_iso == EXPORTED_AT_ISO
    expected_context = {
        **record["daily_context"],
        "tags": tuple(record["daily_context"]["tags"]),
        "coping_used": tuple(record["daily_context"]["coping_used"]),
    }
    assert dict(snapshot.daily_context) == expected_context
    assert dict(snapshot.recording) == record["recording"]
    expected_ids = tuple(question.id for question in DAILY_CORE)
    assert snapshot.answered_field_ids == expected_ids
    assert tuple(field_id for field_id, _ in snapshot.field_status) == tuple(
        question.id for question in (*DAILY_CORE, *DAILY_CONDITIONAL)
    )
    assert [response.field_id for response in snapshot.responses] == [
        question.id for question in (*DAILY_CORE, *DAILY_CONDITIONAL)
    ]
    conditional = snapshot.responses[len(DAILY_CORE) :]
    assert all(response.applicability == "not_applicable" for response in conditional)
    assert all(response.answered is False for response in conditional)
    assert all(response.raw_value is None for response in conditional)
    assert snapshot.responses[0].question_text == DAILY_CORE[0].prompt
    assert snapshot.responses[0].question_kind == "boolean"
    assert snapshot.responses[0].display_value == "没有"
    assert snapshot.visits[0].instrument_id == "daily_nssi_ema"
    assert snapshot.visits[0].instrument_version == "1.0"
    assert snapshot.visits[0].visit_status == "complete"
    assert snapshot.visits[0].instrument_status == "complete"
    assert snapshot.visits[0].completed_at_iso == COMPLETED_AT_ISO
    assert not hasattr(snapshot, "__dict__")
    with pytest.raises(FrozenInstanceError):
        snapshot.visit = "changed"  # type: ignore[misc]


def test_positive_daily_branch_preserves_every_conditional_nssi_answer() -> None:
    record = _completed_record(positive=True)
    snapshot = build_participant_snapshot(
        record,
        visit="daily",
        exported_at_iso=EXPORTED_AT_ISO,
    )

    expected_ids = tuple(
        question.id for question in (*DAILY_CORE, *DAILY_CONDITIONAL)
    )
    assert snapshot.answered_field_ids == expected_ids
    assert tuple(response.field_id for response in snapshot.responses) == expected_ids
    assert all(response.answered for response in snapshot.responses)
    assert all(response.applicability == "applicable" for response in snapshot.responses)
    motive = next(
        response
        for response in snapshot.responses
        if response.field_id == "nssi_motives_24h"
    )
    assert motive.raw_value == DAILY_CONDITIONAL[11].options[:2]
    assert motive.display_value == "、".join(DAILY_CONDITIONAL[11].options[:2])


def test_weekly_day_seven_preserves_protocol_instrument_and_question_order() -> None:
    record = _completed_record(day=7)
    snapshot = build_participant_snapshot(
        record,
        visit="daily",
        exported_at_iso=EXPORTED_AT_ISO,
    )

    expected_instruments = (
        "daily_nssi_ema",
        *(instrument.id for instrument in WEEKLY_INSTRUMENTS),
    )
    assert tuple(visit.instrument_id for visit in snapshot.visits) == expected_instruments
    assert tuple(visit.instrument_version for visit in snapshot.visits) == (
        "1.0",
        *("1.0" for _ in WEEKLY_INSTRUMENTS),
    )
    assert tuple(response.field_id for response in snapshot.responses) == tuple(
        question.id for question in _daily_questions(7)
    )
    weekly_ids = tuple(
        question.id
        for instrument in WEEKLY_INSTRUMENTS
        for question in instrument.questions
    )
    assert snapshot.answered_field_ids[-len(weekly_ids) :] == weekly_ids


@pytest.mark.parametrize("visit", tuple(VISIT_INSTRUMENT_IDS))
def test_each_formal_visit_exports_its_full_ordered_instrument_inventory(
    visit: str,
) -> None:
    record = _completed_record(visit=visit)
    snapshot = build_participant_snapshot(
        record,
        visit=visit,
        exported_at_iso=EXPORTED_AT_ISO,
    )

    expected_questions = _formal_questions(visit)
    assert tuple(item.instrument_id for item in snapshot.visits) == (
        VISIT_INSTRUMENT_IDS[visit]
    )
    assert tuple(response.field_id for response in snapshot.responses) == tuple(
        question.id for question in expected_questions
    )
    assert snapshot.answered_field_ids == tuple(
        question.id for question in expected_questions
    )
    assert all(item.instrument_version == "1.0" for item in snapshot.visits)
    assert all(item.instrument_status == "complete" for item in snapshot.visits)
    assert all(item.completed_at_iso == COMPLETED_AT_ISO for item in snapshot.visits)
    for item, instrument_id in zip(
        snapshot.visits, VISIT_INSTRUMENT_IDS[visit], strict=True
    ):
        expected_ids = tuple(
            question.id for question in FORMAL_INSTRUMENTS[instrument_id].questions
        )
        assert item.answered_field_ids == expected_ids
        assert tuple(field_id for field_id, _ in item.field_status) == expected_ids


@pytest.mark.parametrize("status", ("saved", "skipped", "failed"))
def test_terminal_recording_states_are_sanitized_and_ordered(status: str) -> None:
    record = _completed_record(recording_status=status)
    snapshot = build_participant_snapshot(
        record,
        visit="daily",
        exported_at_iso=EXPORTED_AT_ISO,
    )

    assert tuple(key for key, _ in snapshot.recording) == RECORDING_KEYS
    assert dict(snapshot.recording) == _recording(status)


def test_saved_recording_with_released_devices_builds_valid_export() -> None:
    record = _completed_record()

    bundle = build_participant_export(
        record,
        visit="daily",
        exported_at=EXPORTED_AT,
    )

    assert record["recording"]["saved_confirmed"] is True
    assert record["recording"]["camera_ready"] is False
    assert record["recording"]["microphone_ready"] is False
    assert bundle.data


def test_json_and_workbook_are_canonical_views_of_the_same_snapshot() -> None:
    record = _completed_record(day=7, positive=True)
    snapshot = build_participant_snapshot(
        record,
        visit="daily",
        exported_at_iso=EXPORTED_AT_ISO,
    )
    expected_json = participant_snapshot_json(snapshot)
    expected_sheets = questionnaire_export_sheets(snapshot)

    bundle, names, parsed_json, workbook = _bundle_parts(record, visit="daily")

    assert bundle.filename == "session-20260727-103000.zip"
    assert bundle.mime_type == "application/zip"
    assert names == ["responses.json", "responses.xlsx"]
    assert parsed_json == expected_json
    assert workbook.sheetnames == ["Session", "Responses", "Visits", "Recording"]
    for sheet_name, expected_rows in expected_sheets.items():
        assert _worksheet_rows(workbook[sheet_name]) == list(expected_rows)
    response_headers = list(expected_sheets["Responses"][0])
    assert response_headers == [
        "visit",
        "instrument_id",
        "field_id",
        "question_text",
        "question_kind",
        "answered",
        "applicability",
        "raw_value",
        "display_value",
    ]
    assert isinstance(
        next(
            row["raw_value"]
            for row in expected_sheets["Responses"]
            if row["field_id"] == "nssi_motives_24h"
        ),
        str,
    )


def test_snapshot_deep_copies_source_before_later_mutation() -> None:
    record = _completed_record(positive=True)
    snapshot = build_participant_snapshot(
        record,
        visit="daily",
        exported_at_iso=EXPORTED_AT_ISO,
    )
    before_json = participant_snapshot_json(snapshot)
    before_sheets = questionnaire_export_sheets(snapshot)

    record["daily_context"]["tags"].append("later")
    record["daily_core"]["nssi_urge_now"] = 10
    record["conditional_details"]["nssi_motives_24h"].append("later")
    record["recording"]["status"] = "failed"

    assert participant_snapshot_json(snapshot) == before_json
    assert questionnaire_export_sheets(snapshot) == before_sheets


def test_prohibited_sentinels_never_cross_json_excel_or_archive_boundaries() -> None:
    record = _completed_record(positive=True)
    sentinels = {
        "derived_metrics": "DERIVED_SCORE_51f99",
        "score": "SCORE_32cda",
        "scored_answers": "SCORED_7a992",
        "safety_signals": "SAFETY_f110b",
        "risk": "RISK_7a10c",
        "threshold": "THRESHOLD_d014a",
        "upload": "UPLOAD_01aca",
        "cleanup": "CLEANUP_81bcd",
        "path": "PATH_3e0d9",
        "filename": "FILENAME_8ab2e",
        "device_label": "DEVICE_800a1",
        "media": "MEDIA_091fc",
    }
    record.update(sentinels)
    record["daily_context"]["hostile_extra"] = dict(sentinels)
    record["daily_core"]["hostile_extra"] = dict(sentinels)
    record["recording"].update(sentinels)
    record["formal_visits"]["hostile_extra"] = dict(sentinels)

    bundle, names, parsed_json, workbook = _bundle_parts(record, visit="daily")

    rendered_json = json.dumps(parsed_json, ensure_ascii=False, sort_keys=True)
    workbook_values = "\n".join(
        str(cell.value)
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    boundary_text = "\n".join(
        [rendered_json, workbook_values, bundle.filename, *names]
    )
    assert all(sentinel not in boundary_text for sentinel in sentinels.values())


@pytest.mark.parametrize(
    "mutate",
    [
        lambda record: record.__setitem__("schema_version", 4),
        lambda record: record.__setitem__("schema_version", 5.0),
        lambda record: record.__setitem__("subject_id", 7),
        lambda record: record.__setitem__("record_date", "27-07-2026"),
        lambda record: record.__setitem__("intervention_day", False),
        lambda record: record.__setitem__("visit", "V2"),
        lambda record: record["daily_core"].__setitem__("nssi_urge_now", []),
        lambda record: record["field_status"]["daily"].__setitem__(
            "nssi_urge_now", "scored"
        ),
        lambda record: record["completion"]["answered_field_ids"].__setitem__(
            "daily", ["nssi_urge_now", "nssi_urge_now"]
        ),
        lambda record: record["completion"]["questionnaire_visits"][
            "daily"
        ].__setitem__("completed_at_iso", "private-bad-time"),
        lambda record: record["recording"].__setitem__("version", 1),
        lambda record: record["recording"].__setitem__("version", 2.0),
        lambda record: record["recording"].__setitem__("status", "recording"),
        lambda record: record["recording"].__setitem__(
            "duration_seconds", 2701
        ),
        lambda record: record["recording"].__setitem__("saved_confirmed", False),
    ],
)
def test_malformed_records_fail_closed_with_privacy_safe_errors(mutate) -> None:
    record = _completed_record()
    mutate(record)

    with pytest.raises((TypeError, ValueError)) as captured:
        build_participant_snapshot(
            record,
            visit="daily",
            exported_at_iso=EXPORTED_AT_ISO,
        )

    _assert_no_sentinel_in_exception(captured.value, "private-bad-time")
    assert str(captured.value) in {
        "record is invalid",
        "timestamp is invalid",
        "visit is invalid",
    }


@pytest.mark.parametrize(
    ("visit", "exported_at_iso"),
    [
        ("V1", EXPORTED_AT_ISO),
        ("daily", "2026-07-27T10:30:00"),
        ("daily", "2026-07-27T18:30:00+08:00"),
        ("daily", "2026-07-27T10:30:00.001+00:00"),
    ],
)
def test_snapshot_rejects_wrong_context_or_export_timestamp(
    visit: str, exported_at_iso: str
) -> None:
    with pytest.raises(ValueError):
        build_participant_snapshot(
            _completed_record(),
            visit=visit,
            exported_at_iso=exported_at_iso,
        )


def test_invalid_record_never_reaches_bundle_construction(monkeypatch) -> None:
    record = _completed_record()
    record["recording"]["device_label"] = "PRIVATE_DEVICE_SENTINEL"
    record["recording"]["mode"] = "invalid"
    calls: list[object] = []

    def unexpected_bundle_call(**kwargs):
        calls.append(kwargs)
        raise AssertionError("bundle construction must not be called")

    monkeypatch.setattr(
        questionnaire_export, "build_local_export_bundle", unexpected_bundle_call
    )
    with pytest.raises(ValueError) as captured:
        build_participant_export(
            record,
            visit="daily",
            exported_at=EXPORTED_AT,
        )

    assert calls == []
    _assert_no_sentinel_in_exception(captured.value, "PRIVATE_DEVICE_SENTINEL")


@pytest.mark.parametrize(
    "exported_at",
    [
        datetime(2026, 7, 27, 10, 30),
        datetime(2026, 7, 27, 18, 30, tzinfo=timezone(timedelta(hours=8))),
        datetime(2026, 7, 27, 10, 30, 0, 1, tzinfo=timezone.utc),
    ],
)
def test_export_rejects_invalid_export_datetime_before_bundle_call(
    exported_at: datetime, monkeypatch
) -> None:
    calls: list[object] = []
    monkeypatch.setattr(
        questionnaire_export,
        "build_local_export_bundle",
        lambda **kwargs: calls.append(kwargs),
    )

    with pytest.raises(ValueError):
        build_participant_export(
            _completed_record(),
            visit="daily",
            exported_at=exported_at,
        )

    assert calls == []


def test_snapshot_json_and_sheet_builders_reject_forged_types() -> None:
    with pytest.raises(TypeError):
        participant_snapshot_json({})  # type: ignore[arg-type]
    with pytest.raises(TypeError):
        questionnaire_export_sheets({})  # type: ignore[arg-type]


def _hostile_record(case: str, sentinel: str) -> Mapping[str, object]:
    record = _completed_record(positive=True)
    if case == "top_level_getitem":
        return _ExplodingMapping(
            record,
            sentinel=sentinel,
            operation="getitem",
            key="record_date",
        )
    if case == "nested_mapping_iter":
        record["recording"] = _ExplodingMapping(
            record["recording"], sentinel=sentinel, operation="iter"
        )
        return record
    if case == "nested_mapping_getitem":
        record["recording"] = _ExplodingMapping(
            record["recording"],
            sentinel=sentinel,
            operation="getitem",
            key="version",
        )
        return record
    answered = record["completion"]["answered_field_ids"]["daily"]
    record["completion"]["answered_field_ids"]["daily"] = _ExplodingSequence(
        answered,
        sentinel=sentinel,
        operation="iter" if case == "nested_sequence_iter" else "len",
    )
    return record


@pytest.mark.parametrize(
    "case",
    (
        "top_level_getitem",
        "nested_mapping_iter",
        "nested_mapping_getitem",
        "nested_sequence_iter",
        "nested_sequence_len",
    ),
)
def test_hostile_container_exceptions_are_neutral_at_both_public_boundaries(
    case: str, monkeypatch
) -> None:
    sentinel = f"CALLER_CONTROLLED_{case}_7f1a9"
    record = _hostile_record(case, sentinel)

    with pytest.raises(ValueError, match="^record is invalid$") as snapshot_error:
        build_participant_snapshot(
            record,
            visit="daily",
            exported_at_iso=EXPORTED_AT_ISO,
        )
    _assert_no_sentinel_in_exception(snapshot_error.value, sentinel)
    assert snapshot_error.value.__cause__ is None
    assert snapshot_error.value.__context__ is None

    bundle_calls: list[object] = []
    monkeypatch.setattr(
        questionnaire_export,
        "build_local_export_bundle",
        lambda **kwargs: bundle_calls.append(kwargs),
    )
    with pytest.raises(ValueError, match="^record is invalid$") as export_error:
        build_participant_export(
            record,
            visit="daily",
            exported_at=EXPORTED_AT,
        )
    assert bundle_calls == []
    _assert_no_sentinel_in_exception(export_error.value, sentinel)
    assert export_error.value.__cause__ is None
    assert export_error.value.__context__ is None


@pytest.mark.parametrize(
    ("visit", "exported_at_iso", "message"),
    (
        ("V1", EXPORTED_AT_ISO, "visit is invalid"),
        ("daily", "not-a-timestamp", "timestamp is invalid"),
    ),
)
def test_snapshot_preserves_its_intentional_neutral_error_categories(
    visit: str, exported_at_iso: str, message: str
) -> None:
    with pytest.raises(ValueError, match=f"^{message}$") as captured:
        build_participant_snapshot(
            _completed_record(),
            visit=visit,
            exported_at_iso=exported_at_iso,
        )

    assert captured.value.__cause__ is None
    assert captured.value.__context__ is None


@pytest.mark.parametrize(
    ("sleep_hours", "expected"),
    ((0, 0), (0.0, 0), (0.5, 0.5), (7.0, 7), (7.5, 7.5), (24.0, 24), (24, 24)),
)
def test_sleep_hours_accepts_half_hour_steps_and_canonicalizes_integral_floats(
    sleep_hours, expected
) -> None:
    record = _completed_record()
    record["daily_context"]["sleep_hours"] = sleep_hours

    snapshot = build_participant_snapshot(
        record,
        visit="daily",
        exported_at_iso=EXPORTED_AT_ISO,
    )

    exported = dict(snapshot.daily_context)["sleep_hours"]
    assert exported == expected
    assert type(exported) is type(expected)


@pytest.mark.parametrize(
    "sleep_hours",
    (-0.5, 24.5, -0.0, 7.1, 0.10000000000000002),
)
def test_sleep_hours_rejects_values_just_outside_protocol_range_before_bundle(
    sleep_hours: float, monkeypatch
) -> None:
    record = _completed_record()
    record["daily_context"]["sleep_hours"] = sleep_hours
    bundle_calls: list[object] = []
    monkeypatch.setattr(
        questionnaire_export,
        "build_local_export_bundle",
        lambda **kwargs: bundle_calls.append(kwargs),
    )

    with pytest.raises(ValueError, match="^record is invalid$"):
        build_participant_snapshot(
            record,
            visit="daily",
            exported_at_iso=EXPORTED_AT_ISO,
        )
    with pytest.raises(ValueError, match="^record is invalid$"):
        build_participant_export(
            record,
            visit="daily",
            exported_at=EXPORTED_AT,
        )

    assert bundle_calls == []


def test_daily_context_values_default_sleep_float_exports_as_canonical_int() -> None:
    record = _completed_record()
    record["daily_context"] = {}
    context = daily_context_values(record)
    assert context["sleep_hours"] == 7.0
    assert type(context["sleep_hours"]) is float
    record["daily_context"] = context

    snapshot = build_participant_snapshot(
        record,
        visit="daily",
        exported_at_iso=EXPORTED_AT_ISO,
    )

    exported = dict(snapshot.daily_context)["sleep_hours"]
    assert exported == 7
    assert type(exported) is int


def test_oversized_integer_raw_answer_is_rejected_before_bundle(monkeypatch) -> None:
    record = _completed_record(positive=True)
    record["conditional_details"]["nssi_cut_count_24h"] = 10**100
    bundle_calls: list[object] = []
    monkeypatch.setattr(
        questionnaire_export,
        "build_local_export_bundle",
        lambda **kwargs: bundle_calls.append(kwargs),
    )

    with pytest.raises(ValueError, match="^record is invalid$"):
        build_participant_export(
            record,
            visit="daily",
            exported_at=EXPORTED_AT,
        )

    assert bundle_calls == []


def test_exactly_representable_large_integer_raw_answer_remains_valid() -> None:
    record = _completed_record(positive=True)
    record["conditional_details"]["nssi_cut_count_24h"] = 10**15

    snapshot = build_participant_snapshot(
        record,
        visit="daily",
        exported_at_iso=EXPORTED_AT_ISO,
    )

    response = next(
        item
        for item in snapshot.responses
        if item.field_id == "nssi_cut_count_24h"
    )
    assert response.raw_value == 10**15
    assert type(response.raw_value) is int


def _canonical_cell_json(value: object) -> str:
    return json.dumps(
        value,
        ensure_ascii=False,
        allow_nan=False,
        sort_keys=True,
        separators=(",", ":"),
    )


@pytest.mark.parametrize("extra", (0, 1))
def test_daily_context_cell_uses_final_serialized_excel_length_boundary(
    extra: int, monkeypatch
) -> None:
    empty = _canonical_cell_json({"narrative": ""})
    narrative = "x" * (32_767 - len(empty) + extra)
    record = _completed_record()
    record["daily_context"] = {"narrative": narrative}
    bundle_calls: list[dict[str, object]] = []
    result_sentinel = object()

    def capture_bundle(**kwargs):
        bundle_calls.append(kwargs)
        return result_sentinel

    monkeypatch.setattr(
        questionnaire_export,
        "build_local_export_bundle",
        capture_bundle,
    )
    if extra == 0:
        result = build_participant_export(
            record,
            visit="daily",
            exported_at=EXPORTED_AT,
        )
        assert result is result_sentinel
        assert len(bundle_calls) == 1
        context_cell = bundle_calls[0]["sheets"]["Session"][0]["daily_context"]
        assert len(context_cell) == 32_767
    else:
        with pytest.raises(ValueError, match="^record is invalid$"):
            build_participant_export(
                record,
                visit="daily",
                exported_at=EXPORTED_AT,
            )
        assert bundle_calls == []


@pytest.mark.parametrize("extra", (0, 1))
def test_tuple_raw_cell_uses_final_serialized_excel_length_boundary(
    extra: int, monkeypatch
) -> None:
    empty = _canonical_cell_json([""])
    option = "x" * (32_767 - len(empty) + extra)
    questions = list(questionnaire_export.DAILY_CONDITIONAL)
    index = next(
        index
        for index, question in enumerate(questions)
        if question.id == "nssi_motives_24h"
    )
    questions[index] = replace(questions[index], options=(option,))
    monkeypatch.setattr(
        questionnaire_export, "DAILY_CONDITIONAL", tuple(questions)
    )
    record = _completed_record(positive=True)
    record["conditional_details"]["nssi_motives_24h"] = [option]
    bundle_calls: list[dict[str, object]] = []
    result_sentinel = object()

    def capture_bundle(**kwargs):
        bundle_calls.append(kwargs)
        return result_sentinel

    monkeypatch.setattr(
        questionnaire_export,
        "build_local_export_bundle",
        capture_bundle,
    )
    if extra == 0:
        result = build_participant_export(
            record,
            visit="daily",
            exported_at=EXPORTED_AT,
        )
        assert result is result_sentinel
        assert len(bundle_calls) == 1
        response_rows = bundle_calls[0]["sheets"]["Responses"]
        raw_cell = next(
            row["raw_value"]
            for row in response_rows
            if row["field_id"] == "nssi_motives_24h"
        )
        assert len(raw_cell) == 32_767
    else:
        with pytest.raises(ValueError, match="^record is invalid$"):
            build_participant_export(
                record,
                visit="daily",
                exported_at=EXPORTED_AT,
            )
        assert bundle_calls == []


def _set_exported_text(record: dict[str, object], target: str, value: str) -> None:
    if target == "raw_text":
        record["conditional_details"]["nssi_other_description_24h"] = value
    elif target == "context_text":
        record["daily_context"]["narrative"] = value
    else:
        record["daily_context"]["tags"] = [value]


@pytest.mark.parametrize("target", ("raw_text", "context_text", "context_tuple"))
@pytest.mark.parametrize(
    "bad_text",
    (
        "nul\x00text",
        "carriage\rreturn",
        "noncharacter\ufffe",
        "noncharacter\uffff",
        "ambiguous_xAbC9_escape",
        "unpaired\ud800surrogate",
    ),
)
def test_lossy_exported_text_is_rejected_by_snapshot_before_bundle(
    target: str, bad_text: str, monkeypatch
) -> None:
    record = _completed_record(positive=True)
    _set_exported_text(record, target, bad_text)
    bundle_calls: list[object] = []
    monkeypatch.setattr(
        questionnaire_export,
        "build_local_export_bundle",
        lambda **kwargs: bundle_calls.append(kwargs),
    )

    with pytest.raises(ValueError, match="^record is invalid$"):
        build_participant_snapshot(
            record,
            visit="daily",
            exported_at_iso=EXPORTED_AT_ISO,
        )
    with pytest.raises(ValueError, match="^record is invalid$"):
        build_participant_export(
            record,
            visit="daily",
            exported_at=EXPORTED_AT,
        )

    assert bundle_calls == []


@pytest.mark.parametrize("metadata_field", ("prompt", "low_label"))
def test_lossy_question_metadata_is_rejected_before_serialization(
    metadata_field: str, monkeypatch
) -> None:
    index = 3
    questions = list(questionnaire_export.DAILY_CORE)
    questions[index] = replace(
        questions[index], **{metadata_field: "metadata_x1234_sentinel"}
    )
    monkeypatch.setattr(questionnaire_export, "DAILY_CORE", tuple(questions))

    with pytest.raises(ValueError, match="^record is invalid$"):
        build_participant_snapshot(
            _completed_record(),
            visit="daily",
            exported_at_iso=EXPORTED_AT_ISO,
        )


def test_safe_unicode_newline_and_tab_remain_valid_export_text() -> None:
    record = _completed_record(positive=True)
    safe_text = "安全 Unicode 😀\t标签\n下一行"
    _set_exported_text(record, "raw_text", safe_text)
    _set_exported_text(record, "context_text", safe_text)
    _set_exported_text(record, "context_tuple", safe_text)

    snapshot = build_participant_snapshot(
        record,
        visit="daily",
        exported_at_iso=EXPORTED_AT_ISO,
    )

    response = next(
        item
        for item in snapshot.responses
        if item.field_id == "nssi_other_description_24h"
    )
    assert response.raw_value == safe_text
    assert response.display_value == safe_text
    assert dict(snapshot.daily_context)["narrative"] == safe_text
    assert dict(snapshot.daily_context)["tags"] == (safe_text,)


def test_stateful_recording_mapping_reads_each_allowlisted_key_once_for_real_zip(
) -> None:
    record = _completed_record()
    safe_recording = record["recording"]
    sentinels = {
        key: f"PRIVATE_DEVICE_PATH_SENTINEL_{key}_91ba"
        for key in RECORDING_KEYS
    }
    switching = _SwitchingMapping(
        safe_recording,
        later_values=sentinels,
    )
    record["recording"] = switching

    bundle, names, parsed_json, workbook = _bundle_parts(record, visit="daily")

    assert parsed_json["recording"] == safe_recording
    assert switching.read_counts == {key: 1 for key in RECORDING_KEYS}
    rendered = json.dumps(parsed_json, ensure_ascii=False)
    workbook_values = "\n".join(
        str(cell.value)
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    boundary = "\n".join([bundle.filename, *names, rendered, workbook_values])
    assert all(sentinel not in boundary for sentinel in sentinels.values())


@pytest.mark.parametrize(
    ("day", "visit", "version_key", "instrument_id"),
    (
        (1, "daily", "daily_nssi_ema", "daily_nssi_ema"),
        (7, "daily", "weekly_nssi", "nssi_impulse_weekly"),
        (1, "V1", "formal_nssi_crf", "dshi_lifetime"),
    ),
)
def test_stateful_instrument_versions_are_materialized_once_for_real_zip(
    day: int, visit: str, version_key: str, instrument_id: str
) -> None:
    sentinel = f"PRIVATE_VERSION_SENTINEL_{version_key}_64af"
    record = _completed_record(day=day, visit=visit)
    sentinels = {
        key: f"PRIVATE_VERSION_SENTINEL_{key}_64af"
        for key in INSTRUMENT_VERSION_KEYS
    }
    switching = _SwitchingMapping(
        record["instrument_versions"],
        later_values=sentinels,
    )
    record["instrument_versions"] = switching

    _, _, parsed_json, _ = _bundle_parts(record, visit=visit)

    versions = {
        response["instrument_version"]
        for response in parsed_json["responses"]
        if response["instrument_id"] == instrument_id
    }
    assert versions == {"1.0"}
    assert switching.read_counts == {
        key: 1 for key in INSTRUMENT_VERSION_KEYS
    }
    rendered = json.dumps(parsed_json, ensure_ascii=False)
    assert sentinel not in rendered
    assert all(value not in rendered for value in sentinels.values())


@pytest.mark.parametrize("impostor_kind", ("bool", "subclass"))
@pytest.mark.parametrize(
    "location",
    (
        "top_level_revision",
        "completion_revision",
        "completeness_answered",
        "completeness_required",
    ),
)
def test_integer_metadata_impostors_fail_before_bundle(
    location: str, impostor_kind: str, monkeypatch
) -> None:
    record = _completed_record(visit="V1")
    if location == "top_level_revision":
        target = record
        field = "revision"
    elif location == "completion_revision":
        target = record["completion"]["questionnaire_visits"]["V1"]
        field = "revision"
    else:
        instruments = record["formal_visits"]["V1"]["instruments"]
        payload = next(
            value
            for value in instruments.values()
            if value["completeness"]["answered"] == 1
            and value["completeness"]["required"] == 1
        )
        target = payload["completeness"]
        field = location.removeprefix("completeness_")
    expected = target[field]
    assert type(expected) is int
    target[field] = (
        True
        if impostor_kind == "bool"
        else _EqualitySpoofInt(expected + 1, expected)
    )
    bundle_calls: list[object] = []
    monkeypatch.setattr(
        questionnaire_export,
        "build_local_export_bundle",
        lambda **kwargs: bundle_calls.append(kwargs),
    )

    with pytest.raises(ValueError, match="^record is invalid$"):
        build_participant_export(
            record,
            visit="V1",
            exported_at=EXPORTED_AT,
        )

    assert bundle_calls == []


@pytest.mark.parametrize(
    ("field", "accepted"),
    (
        ("storage", "browser_local"),
        ("status", "saved"),
        ("mode", "long"),
    ),
)
def test_equality_spoofed_recording_enums_fail_before_bundle(
    field: str, accepted: str, monkeypatch
) -> None:
    sentinel = f"INVALID_{field.upper()}_SENTINEL_a810"
    record = _completed_record()
    record["recording"][field] = _EqualitySpoof(sentinel, accepted)
    bundle_calls: list[object] = []
    monkeypatch.setattr(
        questionnaire_export,
        "build_local_export_bundle",
        lambda **kwargs: bundle_calls.append(kwargs),
    )

    with pytest.raises(ValueError, match="^record is invalid$") as captured:
        build_participant_export(
            record,
            visit="daily",
            exported_at=EXPORTED_AT,
        )

    assert bundle_calls == []
    _assert_no_sentinel_in_exception(captured.value, sentinel)


def test_accepted_string_subclasses_are_flattened_in_frozen_snapshot() -> None:
    record = _completed_record(positive=True)
    raw_text = _EqualitySpoof("raw narrative", "raw narrative")
    context_text = _EqualitySpoof("context narrative", "context narrative")
    tag = _EqualitySpoof("home", "home")
    motive = _EqualitySpoof(
        DAILY_CONDITIONAL[11].options[0], DAILY_CONDITIONAL[11].options[0]
    )
    record["conditional_details"]["nssi_other_description_24h"] = raw_text
    record["conditional_details"]["nssi_motives_24h"] = [motive]
    record["daily_context"]["narrative"] = context_text
    record["daily_context"]["tags"] = [tag]

    snapshot = build_participant_snapshot(
        record,
        visit="daily",
        exported_at_iso=EXPORTED_AT_ISO,
    )

    responses = {response.field_id: response for response in snapshot.responses}
    assert type(responses["nssi_other_description_24h"].raw_value) is str
    assert type(responses["nssi_other_description_24h"].display_value) is str
    motives = responses["nssi_motives_24h"].raw_value
    assert type(motives) is tuple
    assert all(type(item) is str for item in motives)
    context = dict(snapshot.daily_context)
    assert type(context["narrative"]) is str
    assert type(context["tags"]) is tuple
    assert all(type(item) is str for item in context["tags"])


def test_formal_payload_version_mismatch_fails_neutrally_before_bundle(
    monkeypatch,
) -> None:
    record = _completed_record(visit="V1")
    record["formal_visits"]["V1"]["instruments"]["dshi_lifetime"][
        "instrument_version"
    ] = "2.0"
    bundle_calls: list[object] = []
    monkeypatch.setattr(
        questionnaire_export,
        "build_local_export_bundle",
        lambda **kwargs: bundle_calls.append(kwargs),
    )

    with pytest.raises(ValueError, match="^record is invalid$"):
        build_participant_export(
            record,
            visit="V1",
            exported_at=EXPORTED_AT,
        )

    assert bundle_calls == []


def test_export_source_has_no_scoring_or_external_capabilities() -> None:
    source = MODULE_SOURCE.read_text(encoding="utf-8")
    tree = ast.parse(source)
    imported_roots = {
        alias.name.split(".")[0]
        for node in ast.walk(tree)
        if isinstance(node, ast.Import)
        for alias in node.names
    }
    imported_modules = {
        (node.module or "").split(".")[0]
        for node in ast.walk(tree)
        if isinstance(node, ast.ImportFrom)
    }
    prohibited_modules = {
        "questionnaire_scoring",
        "os",
        "pathlib",
        "shutil",
        "socket",
        "tempfile",
        "requests",
        "urllib",
        "upload_workflow",
    }
    called_names = {
        node.func.id
        for node in ast.walk(tree)
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Name)
    }

    assert (imported_roots | imported_modules).isdisjoint(prohibited_modules)
    assert called_names.isdisjoint({"open", "Path"})
    assert "questionnaire_scoring" not in source
    assert "derived_metrics" not in source
    assert "score_formal_instrument" not in source

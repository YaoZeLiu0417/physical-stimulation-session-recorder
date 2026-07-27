import ast
from copy import deepcopy
from dataclasses import FrozenInstanceError
from datetime import date, datetime, time, timedelta, timezone, tzinfo
from io import BytesIO
import json
import math
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl
import pytest

import local_export_bundle
from local_export_bundle import LocalExportBundle, build_local_export_bundle


MODULE_SOURCE = Path(__file__).resolve().parents[1] / "local_export_bundle.py"
EXPORTED_AT = datetime(2026, 7, 27, 10, 30, tzinfo=timezone.utc)
ZIP_MEMBERS = ["responses.json", "responses.xlsx"]


def _build_bundle(
    *,
    snapshot: dict[str, object] | None = None,
    sheets: dict[str, list[dict[str, object]]] | None = None,
    exported_at: datetime = EXPORTED_AT,
    filename_prefix: str = "session",
) -> LocalExportBundle:
    return build_local_export_bundle(
        snapshot={"schema_version": 1} if snapshot is None else snapshot,
        sheets={"Session": [{"field": "value"}]} if sheets is None else sheets,
        exported_at=exported_at,
        filename_prefix=filename_prefix,
    )


def _read_bundle(bundle: LocalExportBundle) -> tuple[dict[str, object], bytes]:
    with ZipFile(BytesIO(bundle.data), "r") as archive:
        return (
            json.loads(archive.read("responses.json")),
            archive.read("responses.xlsx"),
        )


def _load_workbook(bundle: LocalExportBundle) -> openpyxl.Workbook:
    _, workbook_bytes = _read_bundle(bundle)
    return openpyxl.load_workbook(BytesIO(workbook_bytes), data_only=False)


def test_bundle_has_exact_neutral_filename_mime_type_and_zip_member_order() -> None:
    bundle = _build_bundle(
        snapshot={"participant_id": "sub-001"},
        sheets={"Session": [{"participant_id": "sub-001"}]},
    )

    assert bundle.filename == "session-20260727-103000.zip"
    assert "sub-001" not in bundle.filename
    assert bundle.mime_type == "application/zip"
    assert isinstance(bundle.data, bytes)
    assert not hasattr(bundle, "__dict__")
    with pytest.raises(FrozenInstanceError):
        bundle.filename = "changed.zip"  # type: ignore[misc]

    with ZipFile(BytesIO(bundle.data), "r") as archive:
        assert archive.namelist() == ZIP_MEMBERS
        assert [info.is_dir() for info in archive.infolist()] == [False, False]
        assert [info.compress_type for info in archive.infolist()] == [
            ZIP_DEFLATED,
            ZIP_DEFLATED,
        ]
        assert [info.date_time for info in archive.infolist()] == [
            (1980, 1, 1, 0, 0, 0),
            (1980, 1, 1, 0, 0, 0),
        ]


def test_json_is_canonical_utf8_and_preserves_raw_values() -> None:
    snapshot = {
        "unicode": "中文\t=raw\ntext",
        "nested": {
            "zero": 0,
            "false": False,
            "none": None,
            "list": ["é", 0, False, None, {"z": 2, "a": 1}],
        },
        "alpha": "first after sorting",
    }
    expected = json.dumps(
        snapshot,
        ensure_ascii=False,
        allow_nan=False,
        sort_keys=True,
        indent=2,
        separators=(",", ": "),
    ).encode("utf-8")

    bundle = _build_bundle(snapshot=snapshot)

    with ZipFile(BytesIO(bundle.data), "r") as archive:
        encoded = archive.read("responses.json")
    assert encoded == expected
    assert json.loads(encoded) == snapshot
    assert b"\\u4e2d" not in encoded


def test_generation_is_binary_deterministic_for_equal_canonical_inputs() -> None:
    snapshot_a = {"z": 3, "a": {"right": False, "left": 0}}
    snapshot_b = {"a": {"left": 0, "right": False}, "z": 3}
    sheets = {
        "Session": [{"key": "value", "count": 0}],
        "Responses": [{"raw": ["a", False, None]}],
    }

    first = _build_bundle(snapshot=snapshot_a, sheets=sheets)
    second = _build_bundle(snapshot=snapshot_b, sheets=deepcopy(sheets))

    assert first == second
    assert first.data == second.data


def test_workbook_preserves_sheet_and_first_seen_header_order_and_values() -> None:
    sheets = {
        "Session": [
            {
                "visit": "daily",
                "subject": "参与者",
                "zero": 0,
                "false": False,
                "none": None,
                "nested": {"z": 2, "a": 1},
                "items": ["α", 0, False, None],
            },
            {"late": "second-row", "visit": "followup"},
        ],
        "Responses": [{"item_id": "q1", "raw": "原始值"}],
    }

    workbook = _load_workbook(_build_bundle(sheets=sheets))

    assert workbook.sheetnames == ["Session", "Responses"]
    worksheet = workbook["Session"]
    assert [cell.value for cell in worksheet[1]] == [
        "visit",
        "subject",
        "zero",
        "false",
        "none",
        "nested",
        "items",
        "late",
    ]
    assert [cell.value for cell in worksheet[2]] == [
        "daily",
        "参与者",
        0,
        False,
        None,
        '{"a":1,"z":2}',
        '["α",0,false,null]',
        None,
    ]
    assert [cell.value for cell in worksheet[3]] == [
        "followup",
        None,
        None,
        None,
        None,
        None,
        None,
        "second-row",
    ]


def test_empty_worksheets_and_empty_rows_remain_visible_and_structured() -> None:
    workbook = _load_workbook(
        _build_bundle(sheets={"NoRows": [], "EmptyRow": [{}]})
    )

    assert workbook.sheetnames == ["NoRows", "EmptyRow"]
    for worksheet in workbook.worksheets:
        assert worksheet.sheet_state == "visible"
        assert worksheet.freeze_panes == "A2"
        assert worksheet.auto_filter.ref == "A1"
        assert 12 <= worksheet.column_dimensions["A"].width <= 48
    assert 18 <= workbook["EmptyRow"].row_dimensions[2].height <= 90


def test_workbook_has_exact_restrained_style_and_bounded_dimensions() -> None:
    bundle = _build_bundle(
        sheets={
            "Responses": [
                {
                    "item_id": "q1",
                    "answer": "wrapped text " * 25,
                    "number": 12.5,
                },
                {"item_id": "q2", "answer": "line one\nline two"},
            ]
        }
    )
    worksheet = _load_workbook(bundle)["Responses"]

    assert worksheet.sheet_state == "visible"
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref == "A1:C3"
    assert worksheet.sheet_properties.tabColor.rgb == "FF2D2674"
    for cell in worksheet[1]:
        assert cell.fill.fill_type == "solid"
        assert cell.fill.fgColor.rgb == "FF000035"
        assert cell.font.bold is True
        assert cell.font.color.rgb == "FFFFFFFF"
        assert cell.alignment.wrap_text is True
        assert cell.alignment.vertical == "top"
    assert worksheet["A2"].font.bold is True
    assert worksheet["A2"].font.color.rgb == "FF2D2674"
    assert worksheet["B2"].alignment.wrap_text is True
    assert worksheet["B2"].alignment.vertical == "top"
    for dimension in worksheet.column_dimensions.values():
        assert 12 <= dimension.width <= 48
    for row_index in range(1, worksheet.max_row + 1):
        height = worksheet.row_dimensions[row_index].height
        assert height is not None
        assert 18 <= height <= 90


def test_workbook_writes_approved_date_like_cells_as_typed_values() -> None:
    aware = datetime(2026, 7, 27, 18, 30, tzinfo=timezone(timedelta(hours=8)))
    workbook = _load_workbook(
        _build_bundle(
            sheets={
                "Typed": [
                    {
                        "date": date(2026, 7, 27),
                        "datetime": aware,
                        "time": time(10, 30, 15),
                        "duration": timedelta(hours=1, minutes=2, seconds=3),
                    }
                ]
            }
        )
    )
    worksheet = workbook["Typed"]

    assert worksheet["A2"].value == datetime(2026, 7, 27)
    assert worksheet["B2"].value == datetime(2026, 7, 27, 10, 30)
    assert worksheet["C2"].value == time(10, 30, 15)
    assert worksheet["D2"].value == timedelta(hours=1, minutes=2, seconds=3)


def test_excel_date_and_datetime_boundaries_round_trip_exactly() -> None:
    maximum_datetime = datetime(9999, 12, 31, 23, 59, 59, 999_000)
    values = {
        "minimum_date": date(1900, 1, 1),
        "modern_date": date(2026, 7, 27),
        "maximum_date": date.max,
        "minimum_datetime": datetime(1900, 1, 1),
        "minimum_datetime_with_time": datetime(1900, 1, 1, 12, 30, 45, 123_000),
        "modern_datetime": datetime(2026, 7, 27, 10, 30, 15, 123_000),
        "maximum_datetime": maximum_datetime,
        "normalized_minimum": datetime(
            1900,
            1,
            1,
            8,
            tzinfo=timezone(timedelta(hours=8)),
        ),
    }

    worksheet = _load_workbook(
        _build_bundle(sheets={"Dates": [values]})
    )["Dates"]

    assert [cell.value for cell in worksheet[2]] == [
        datetime(1900, 1, 1),
        datetime(2026, 7, 27),
        datetime(9999, 12, 31),
        datetime(1900, 1, 1),
        datetime(1900, 1, 1, 12, 30, 45, 123_000),
        datetime(2026, 7, 27, 10, 30, 15, 123_000),
        maximum_datetime,
        datetime(1900, 1, 1),
    ]


@pytest.mark.parametrize(
    "value",
    [
        date.min,
        date(1899, 12, 31),
        datetime.min,
        datetime(1899, 12, 31, 23, 59, 59, 999_000),
        datetime(2026, 7, 27, 10, 30, 15, 123_456),
        datetime.max,
    ],
)
def test_unrepresentable_or_submillisecond_dates_fail_before_workbook_creation(
    value: object,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def fail_workbook(*args: object, **kwargs: object) -> object:
        raise AssertionError("workbook creation must not be reached")

    monkeypatch.setattr(local_export_bundle.xlsxwriter, "Workbook", fail_workbook)

    with pytest.raises(ValueError, match="Excel|millisecond"):
        _build_bundle(sheets={"Dates": [{"value": value}]})


class _FailingTimezone(tzinfo):
    def utcoffset(self, dt: datetime | None) -> timedelta:
        raise RuntimeError("private timezone failure")

    def dst(self, dt: datetime | None) -> timedelta:
        return timedelta(0)


@pytest.mark.parametrize(
    "value",
    [
        datetime(1, 1, 1, tzinfo=timezone(timedelta(hours=14))),
        datetime.max.replace(tzinfo=timezone(timedelta(hours=-14))),
        datetime(2026, 7, 27, tzinfo=_FailingTimezone()),
    ],
)
def test_timezone_normalization_failures_are_sanitized_before_workbook_creation(
    value: datetime,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def fail_workbook(*args: object, **kwargs: object) -> object:
        raise AssertionError("workbook creation must not be reached")

    monkeypatch.setattr(local_export_bundle.xlsxwriter, "Workbook", fail_workbook)

    with pytest.raises(ValueError, match="datetime") as error:
        _build_bundle(sheets={"Dates": [{"value": value}]})
    assert "private timezone failure" not in str(error.value)


def test_time_and_timedelta_representable_boundaries_round_trip_exactly() -> None:
    maximum_duration = timedelta(
        days=2_958_465,
        seconds=86_399,
        microseconds=999_000,
    )
    values = {
        "minimum_time": time.min,
        "modern_time": time(10, 30, 15, 123_000),
        "maximum_time": time(23, 59, 59, 999_000),
        "minimum_duration": -maximum_duration,
        "negative_duration": timedelta(milliseconds=-1),
        "zero_duration": timedelta(0),
        "modern_duration": timedelta(days=30, milliseconds=123),
        "maximum_duration": maximum_duration,
    }

    worksheet = _load_workbook(
        _build_bundle(sheets={"Dates": [values]})
    )["Dates"]

    assert [cell.value for cell in worksheet[2]] == list(values.values())
    assert [type(cell.value) for cell in worksheet[2]] == [
        time,
        time,
        time,
        timedelta,
        timedelta,
        timedelta,
        timedelta,
        timedelta,
    ]


@pytest.mark.parametrize(
    "value",
    [
        time(10, 30, 15, 123_456),
        time.max,
        timedelta(microseconds=1),
        timedelta(microseconds=-1),
        timedelta(
            days=2_958_466,
        ),
        timedelta(days=-2_958_466),
        timedelta.max,
    ],
)
def test_time_and_timedelta_values_that_cannot_round_trip_are_rejected(
    value: object,
) -> None:
    with pytest.raises(ValueError, match="Excel|millisecond"):
        _build_bundle(sheets={"Dates": [{"value": value}]})


def test_formula_and_url_like_text_stays_plain_and_json_stays_raw() -> None:
    raw_values = [
        "=1+1",
        "+SUM(A1:A2)",
        "-1+2",
        "@cmd",
        "https://example.invalid/path",
        "www.example.invalid",
        "\t=not-leading",
        "normal",
    ]
    bundle = _build_bundle(
        snapshot={"raw": raw_values},
        sheets={"Responses": [{"value": value} for value in raw_values]},
    )

    parsed_json, _ = _read_bundle(bundle)
    assert parsed_json == {"raw": raw_values}
    worksheet = _load_workbook(bundle)["Responses"]
    expected_cells = [
        "'=1+1",
        "'+SUM(A1:A2)",
        "'-1+2",
        "'@cmd",
        "https://example.invalid/path",
        "www.example.invalid",
        "\t=not-leading",
        "normal",
    ]
    assert [worksheet.cell(row=index, column=1).value for index in range(2, 10)] == expected_cells
    for row in worksheet.iter_rows():
        for cell in row:
            assert cell.data_type != "f"
            assert cell.hyperlink is None


@pytest.mark.parametrize(
    "name",
    [
        "",
        "a" * 32,
        "bad/name",
        "bad\\name",
        "bad:name",
        "bad*name",
        "bad?name",
        "bad[name",
        "bad]name",
        "bad\x00name",
        "bad\x1fname",
        "'leading",
        "trailing'",
    ],
)
def test_invalid_sheet_names_are_rejected(name: str) -> None:
    with pytest.raises(ValueError, match="sheet name"):
        _build_bundle(sheets={name: []})


@pytest.mark.parametrize(
    "sheets",
    [
        {"Data": [], "data": []},
        {"A": [], "Ａ": []},
    ],
)
def test_case_insensitive_normalized_duplicate_sheet_names_are_rejected(
    sheets: dict[str, list[dict[str, object]]],
) -> None:
    with pytest.raises(ValueError, match="duplicate sheet name"):
        _build_bundle(sheets=sheets)


@pytest.mark.parametrize(
    "prefix",
    [
        "",
        "-session",
        "session-",
        "session--copy",
        "Session",
        "session_name",
        "session.name",
        "session/name",
        "session\\name",
        " session",
        "session ",
        "会话",
        "a" * 65,
        "participant-sub-001",
        "study-tavns",
        "intervention-day-7",
        "export",
    ],
)
def test_unsafe_filename_prefixes_are_rejected(prefix: str) -> None:
    with pytest.raises(ValueError, match="filename_prefix"):
        _build_bundle(filename_prefix=prefix)


def test_safe_custom_filename_prefix_is_supported() -> None:
    bundle = _build_bundle(filename_prefix="synthetic-session")

    assert bundle.filename == "synthetic-session-20260727-103000.zip"


def test_empty_containers_consume_one_shared_cumulative_budget(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(local_export_bundle, "_MAX_TOTAL_VALUES", 6)

    _build_bundle(snapshot={}, sheets={"Data": [{}, {}]})
    with pytest.raises(ValueError, match="too many values"):
        _build_bundle(snapshot={}, sheets={"Data": [{}, {}, {}]})


def test_empty_worksheets_consume_the_shared_cumulative_budget(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(local_export_bundle, "_MAX_TOTAL_VALUES", 4)
    _build_bundle(snapshot={}, sheets={"Only": []})

    monkeypatch.setattr(local_export_bundle, "_MAX_TOTAL_VALUES", 5)
    with pytest.raises(ValueError, match="too many values"):
        _build_bundle(snapshot={}, sheets={"First": [], "Second": []})


@pytest.mark.parametrize(
    "exported_at",
    [
        datetime(2026, 7, 27, 10, 30),
        datetime(2026, 7, 27, 10, 30, tzinfo=timezone(timedelta(hours=8))),
        datetime(2026, 7, 27, 10, 30, 0, 1, tzinfo=timezone.utc),
    ],
)
def test_export_time_must_be_utc_aware_and_second_precision(
    exported_at: datetime,
) -> None:
    with pytest.raises(ValueError, match="exported_at"):
        _build_bundle(exported_at=exported_at)


@pytest.mark.parametrize("bad_number", [math.nan, math.inf, -math.inf])
@pytest.mark.parametrize("location", ["snapshot", "sheet"])
def test_non_finite_numbers_are_rejected(bad_number: float, location: str) -> None:
    kwargs: dict[str, object] = {}
    if location == "snapshot":
        kwargs["snapshot"] = {"bad": [bad_number]}
    else:
        kwargs["sheets"] = {"Data": [{"bad": {"number": bad_number}}]}

    with pytest.raises(ValueError, match="finite"):
        _build_bundle(**kwargs)  # type: ignore[arg-type]


@pytest.mark.parametrize(
    ("snapshot", "sheets"),
    [
        ({"bad": object()}, {"Data": []}),
        ({"bad": (1, 2)}, {"Data": []}),
        ({1: "bad"}, {"Data": []}),
        ({}, {"Data": [{1: "bad"}]}),
        ({}, {1: []}),
        ({}, {"Data": [object()]}),
    ],
)
def test_unsupported_values_non_string_keys_and_invalid_rows_are_rejected(
    snapshot: dict[object, object],
    sheets: dict[object, object],
) -> None:
    with pytest.raises((TypeError, ValueError)):
        _build_bundle(snapshot=snapshot, sheets=sheets)  # type: ignore[arg-type]


def test_cycles_and_excessive_nesting_are_rejected() -> None:
    cyclic_snapshot: dict[str, object] = {}
    cyclic_snapshot["self"] = cyclic_snapshot
    cyclic_cell: list[object] = []
    cyclic_cell.append(cyclic_cell)
    too_deep: object = "end"
    for _ in range(70):
        too_deep = [too_deep]

    with pytest.raises(ValueError, match="cycle"):
        _build_bundle(snapshot=cyclic_snapshot)
    with pytest.raises(ValueError, match="cycle"):
        _build_bundle(sheets={"Data": [{"value": cyclic_cell}]})
    with pytest.raises(ValueError, match="deep"):
        _build_bundle(snapshot={"value": too_deep})


def test_overlong_excel_text_is_rejected_without_truncation() -> None:
    with pytest.raises(ValueError, match="32,767"):
        _build_bundle(sheets={"Data": [{"value": "x" * 32_768}]})
    with pytest.raises(ValueError, match="32,767"):
        _build_bundle(sheets={"Data": [{"value": "=" + "x" * 32_766}]})


def test_inputs_are_not_mutated_by_successful_generation() -> None:
    snapshot = {"nested": {"list": ["=raw", False, 0, None]}}
    sheets = {
        "Data": [
            {
                "key": "=formula-risk",
                "nested": {"z": [1, 2], "a": False},
            }
        ]
    }
    original_snapshot = deepcopy(snapshot)
    original_sheets = deepcopy(sheets)

    _build_bundle(snapshot=snapshot, sheets=sheets)

    assert snapshot == original_snapshot
    assert sheets == original_sheets


def test_generation_failure_leaves_caller_data_unchanged(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    snapshot = {"nested": {"list": ["raw", False, 0, None]}}
    sheets = {"Data": [{"key": "value", "nested": {"a": [1, 2]}}]}
    original_snapshot = deepcopy(snapshot)
    original_sheets = deepcopy(sheets)

    def fail_workbook(*args: object, **kwargs: object) -> object:
        raise RuntimeError("synthetic workbook failure")

    monkeypatch.setattr(local_export_bundle.xlsxwriter, "Workbook", fail_workbook)

    with pytest.raises(RuntimeError, match="synthetic workbook failure"):
        _build_bundle(snapshot=snapshot, sheets=sheets)
    assert snapshot == original_snapshot
    assert sheets == original_sheets


def test_source_boundary_is_memory_only_and_has_no_external_capability() -> None:
    source = MODULE_SOURCE.read_text(encoding="utf-8")
    source_lower = source.lower()
    tree = ast.parse(source)
    imported_roots = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            imported_roots.update(alias.name.split(".", 1)[0] for alias in node.names)
        elif isinstance(node, ast.ImportFrom) and node.module:
            imported_roots.add(node.module.split(".", 1)[0])

    assert imported_roots <= {
        "collections",
        "dataclasses",
        "datetime",
        "io",
        "json",
        "math",
        "re",
        "unicodedata",
        "zipfile",
        "xlsxwriter",
    }
    assert not {
        "os",
        "pathlib",
        "requests",
        "shutil",
        "socket",
        "streamlit",
        "tempfile",
        "urllib",
    } & imported_roots
    prohibited_fragments = {
        "dailyrecordstore",
        "mediarecorder",
        "questionnaire",
        "record_store",
        "streamlit",
        "tempfile",
        "upload_",
        "webrtc",
    }
    assert [value for value in prohibited_fragments if value in source_lower] == []
    prohibited_calls = {
        "open",
        "read_bytes",
        "write_bytes",
        "read_text",
        "write_text",
        "unlink",
        "mkdir",
        "makedirs",
        "request",
        "urlopen",
        "upload",
    }
    calls = []
    for node in ast.walk(tree):
        if not isinstance(node, ast.Call):
            continue
        if isinstance(node.func, ast.Name):
            calls.append(node.func.id.lower())
        elif isinstance(node.func, ast.Attribute):
            calls.append(node.func.attr.lower())
    assert sorted(set(calls) & prohibited_calls) == []

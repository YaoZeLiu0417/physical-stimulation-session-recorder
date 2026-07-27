import ast
from dataclasses import FrozenInstanceError
from datetime import datetime, timedelta, timezone, tzinfo
from io import BytesIO
import json
from pathlib import Path
import xml.etree.ElementTree as ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl
import pytest

import showcase_export
from showcase_audit import FORBIDDEN_TERMS
from showcase_export import SyntheticShowcaseArchive, build_synthetic_showcase_zip


MODULE_SOURCE = Path(__file__).resolve().parents[1] / "showcase_export.py"
GENERATED_AT = datetime(2026, 7, 28, 9, 15, 30, tzinfo=timezone.utc)
ITEM_IDS = (
    "demo_process_clarity",
    "demo_camera_smoothness",
    "demo_information_load",
    "demo_workflow_willingness",
)
ZIP_MEMBERS = ["responses.json", "responses.xlsx"]
PROHIBITED_CONTENT_TERMS = FORBIDDEN_TERMS + (
    "questionnaire",
    "research",
    "study",
    "score",
    "risk",
    "participant",
    "subject",
    "path",
    "upload",
)
OOXML_RAW_PROHIBITED_TERMS = tuple(
    term for term in PROHIBITED_CONTENT_TERMS if term not in {"path", "subject"}
)
SAFE_IMPORT_ROOTS = {
    "__future__",
    "dataclasses",
    "datetime",
    "local_export_bundle",
}


class _HostileTimezone(tzinfo):
    def utcoffset(self, value: datetime | None) -> timedelta | None:
        raise RuntimeError("PRIVATE_GENERATION_TIMEZONE_7F3A")

    def dst(self, value: datetime | None) -> timedelta | None:
        return None


class _MutableTimezone(tzinfo):
    def __init__(self) -> None:
        self.utcoffset_calls = 0

    def utcoffset(self, value: datetime | None) -> timedelta | None:
        self.utcoffset_calls += 1
        if self.utcoffset_calls == 1:
            return timedelta(0)
        return timedelta(hours=9)

    def dst(self, value: datetime | None) -> timedelta | None:
        return None


class _HostileDatetime(datetime):
    def isoformat(self, *args: object, **kwargs: object) -> str:
        raise RuntimeError("PRIVATE_DATETIME_ABC123")


class _RaisingOffset(timedelta):
    def __new__(cls) -> "_RaisingOffset":
        return super().__new__(cls)

    def __ne__(self, other: object) -> bool:
        raise RuntimeError("PRIVATE_OFFSET_COMPARE_77")


class _LyingNonzeroOffset(timedelta):
    def __new__(cls) -> "_LyingNonzeroOffset":
        return super().__new__(cls, hours=9)

    def __ne__(self, other: object) -> bool:
        return False


class _OffsetTimezone(tzinfo):
    def __init__(self, offset: timedelta) -> None:
        self.offset = offset
        self.utcoffset_calls = 0

    def utcoffset(self, value: datetime | None) -> timedelta | None:
        self.utcoffset_calls += 1
        return self.offset

    def dst(self, value: datetime | None) -> timedelta | None:
        return None


def _build_archive(
    *,
    process_clarity: int = 4,
    camera_smoothness: int | None = 3,
    information_load: int = 2,
    workflow_willingness: int = 1,
    recording_state: str = "saved",
    generated_at: datetime = GENERATED_AT,
) -> SyntheticShowcaseArchive:
    return build_synthetic_showcase_zip(
        process_clarity=process_clarity,
        camera_smoothness=camera_smoothness,
        information_load=information_load,
        workflow_willingness=workflow_willingness,
        recording_state=recording_state,
        generated_at=generated_at,
    )


def _read_archive(
    archive: SyntheticShowcaseArchive,
) -> tuple[dict[str, object], openpyxl.Workbook]:
    with ZipFile(BytesIO(archive.data), "r") as zip_file:
        assert zip_file.namelist() == ZIP_MEMBERS
        snapshot = json.loads(zip_file.read("responses.json"))
        workbook_bytes = zip_file.read("responses.xlsx")
    workbook = openpyxl.load_workbook(BytesIO(workbook_bytes), data_only=True)
    return snapshot, workbook


def _sheet_rows(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
) -> list[dict[str, object]]:
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows
    headers = rows[0]
    return [dict(zip(headers, values, strict=True)) for values in rows[1:]]


def _archive_member_bytes(
    archive: SyntheticShowcaseArchive,
    member_name: str,
) -> bytes:
    with ZipFile(BytesIO(archive.data), "r") as zip_file:
        return zip_file.read(member_name)


def _replace_archive_member(
    archive: SyntheticShowcaseArchive,
    member_name: str,
    replacement: bytes,
) -> SyntheticShowcaseArchive:
    output = BytesIO()
    with ZipFile(BytesIO(archive.data), "r") as source:
        with ZipFile(output, "w", compression=ZIP_DEFLATED) as target:
            for member in source.infolist():
                payload = (
                    replacement
                    if member.filename == member_name
                    else source.read(member.filename)
                )
                target.writestr(member, payload)
    return SyntheticShowcaseArchive(filename=archive.filename, data=output.getvalue())


def _add_workbook_xml_payload(
    workbook_bytes: bytes,
    member_name: str,
    payload: bytes,
) -> bytes:
    output = BytesIO()
    with ZipFile(BytesIO(workbook_bytes), "r") as source:
        with ZipFile(output, "w", compression=ZIP_DEFLATED) as target:
            for member in source.infolist():
                target.writestr(member, source.read(member.filename))
            target.writestr(member_name, payload)
    return output.getvalue()


def _assert_source_imports_are_isolated(source: str) -> None:
    tree = ast.parse(source)
    imported_roots: set[str] = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            imported_roots.update(alias.name.split(".", 1)[0] for alias in node.names)
        elif isinstance(node, ast.ImportFrom) and node.module:
            imported_roots.add(node.module.split(".", 1)[0])
    assert imported_roots == SAFE_IMPORT_ROOTS


def _assert_text_has_no_private_content(
    text: str,
    *,
    location: str,
    terms: tuple[str, ...] = PROHIBITED_CONTENT_TERMS,
) -> None:
    folded_text = text.casefold()
    for term in terms:
        assert term.casefold() not in folded_text, f"{term}: {location}"


def _assert_bytes_have_no_private_content(
    payload: bytes,
    *,
    location: str,
    terms: tuple[str, ...] = PROHIBITED_CONTENT_TERMS,
) -> None:
    folded_payload = payload.lower()
    for term in terms:
        assert term.casefold().encode("utf-8") not in folded_payload, (
            f"{term}: {location}"
        )


def _xml_content_values(payload: bytes, *, location: str) -> list[str]:
    try:
        root = ElementTree.fromstring(payload)
    except ElementTree.ParseError:
        raise AssertionError(f"invalid XML payload: {location}") from None

    values: list[str] = []
    for element in root.iter():
        if element.text:
            values.append(element.text)
        if element.tail:
            values.append(element.tail)
        values.extend(element.attrib.values())
    return values


def _assert_archive_has_no_private_content(
    archive: SyntheticShowcaseArchive,
) -> None:
    _assert_text_has_no_private_content(
        archive.filename,
        location="archive filename",
    )
    with ZipFile(BytesIO(archive.data), "r") as outer_archive:
        assert outer_archive.namelist() == ZIP_MEMBERS
        json_bytes = outer_archive.read("responses.json")
        workbook_bytes = outer_archive.read("responses.xlsx")

    _assert_bytes_have_no_private_content(
        json_bytes,
        location="responses.json bytes",
    )
    try:
        json_text = json_bytes.decode("utf-8")
    except UnicodeDecodeError:
        raise AssertionError("responses.json is not valid UTF-8") from None
    _assert_text_has_no_private_content(
        json_text,
        location="responses.json text",
    )
    snapshot = json.loads(json_text)
    _assert_text_has_no_private_content(
        json.dumps(snapshot, ensure_ascii=False),
        location="parsed responses.json",
    )

    with ZipFile(BytesIO(workbook_bytes), "r") as workbook_package:
        for member in workbook_package.infolist():
            if member.is_dir() or not member.filename.casefold().endswith(
                (".xml", ".rels")
            ):
                continue
            _assert_text_has_no_private_content(
                member.filename,
                location="workbook package member name",
            )
            payload = workbook_package.read(member)
            _assert_bytes_have_no_private_content(
                payload,
                location=member.filename,
                terms=OOXML_RAW_PROHIBITED_TERMS,
            )
            try:
                package_text = payload.decode("utf-8-sig")
            except UnicodeDecodeError:
                raise AssertionError(
                    f"workbook text payload is not UTF-8: {member.filename}"
                ) from None
            _assert_text_has_no_private_content(
                package_text,
                location=member.filename,
                terms=OOXML_RAW_PROHIBITED_TERMS,
            )
            _assert_text_has_no_private_content(
                "\n".join(
                    _xml_content_values(payload, location=member.filename)
                ),
                location=f"{member.filename} structured content",
            )

    workbook = openpyxl.load_workbook(BytesIO(workbook_bytes), data_only=True)
    searchable_values: list[str] = []
    for worksheet in workbook.worksheets:
        searchable_values.append(worksheet.title)
        searchable_values.extend(
            str(cell.value)
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None
        )
    _assert_text_has_no_private_content(
        "\n".join(searchable_values),
        location="workbook sheets and cells",
    )


def test_saved_archive_has_exact_public_shape_and_matching_values() -> None:
    archive = _build_archive()

    assert archive.filename == "synthetic-session-20260728-091530.zip"
    assert isinstance(archive.data, bytes)
    assert not hasattr(archive, "__dict__")
    with pytest.raises(FrozenInstanceError):
        archive.filename = "changed.zip"  # type: ignore[misc]

    snapshot, workbook = _read_archive(archive)
    assert snapshot == {
        "generated_at_utc": "2026-07-28T09:15:30Z",
        "ratings": [
            {
                "applicable": True,
                "item_id": "demo_process_clarity",
                "value": 4,
            },
            {
                "applicable": True,
                "item_id": "demo_camera_smoothness",
                "value": 3,
            },
            {
                "applicable": True,
                "item_id": "demo_information_load",
                "value": 2,
            },
            {
                "applicable": True,
                "item_id": "demo_workflow_willingness",
                "value": 1,
            },
        ],
        "recording": {"state": "saved", "synthetic": True},
        "schema_version": 1,
    }
    assert workbook.sheetnames == ["Session", "Responses", "Recording"]
    assert _sheet_rows(workbook["Session"]) == [
        {
            "schema_version": 1,
            "generated_at_utc": "2026-07-28T09:15:30Z",
        }
    ]
    assert _sheet_rows(workbook["Responses"]) == snapshot["ratings"]
    assert _sheet_rows(workbook["Recording"]) == [snapshot["recording"]]


@pytest.mark.parametrize("recording_state", ["skipped", "failed"])
def test_no_camera_rating_is_explicitly_not_applicable_in_both_formats(
    recording_state: str,
) -> None:
    archive = _build_archive(
        camera_smoothness=None,
        recording_state=recording_state,
    )

    snapshot, workbook = _read_archive(archive)
    camera_json = snapshot["ratings"][1]  # type: ignore[index]
    camera_xlsx = _sheet_rows(workbook["Responses"])[1]

    assert camera_json == {
        "applicable": False,
        "item_id": "demo_camera_smoothness",
        "value": None,
    }
    assert camera_xlsx == camera_json
    assert _sheet_rows(workbook["Recording"]) == [
        {"state": recording_state, "synthetic": True}
    ]
    assert snapshot["recording"] == {
        "state": recording_state,
        "synthetic": True,
    }


@pytest.mark.parametrize("recording_state", ["skipped", "failed"])
def test_non_saved_recording_may_retain_an_explicit_camera_rating(
    recording_state: str,
) -> None:
    snapshot, workbook = _read_archive(
        _build_archive(camera_smoothness=0, recording_state=recording_state)
    )

    assert snapshot["ratings"][1] == {  # type: ignore[index]
        "applicable": True,
        "item_id": "demo_camera_smoothness",
        "value": 0,
    }
    assert _sheet_rows(workbook["Responses"])[1] == (
        snapshot["ratings"][1]  # type: ignore[index]
    )


@pytest.mark.parametrize(
    ("field", "value", "error_type"),
    [
        ("process_clarity", True, TypeError),
        ("process_clarity", 1.0, TypeError),
        ("process_clarity", -1, ValueError),
        ("process_clarity", 5, ValueError),
        ("camera_smoothness", False, TypeError),
        ("camera_smoothness", 2.0, TypeError),
        ("camera_smoothness", -1, ValueError),
        ("camera_smoothness", 5, ValueError),
        ("information_load", True, TypeError),
        ("information_load", 1.0, TypeError),
        ("information_load", -1, ValueError),
        ("information_load", 5, ValueError),
        ("workflow_willingness", True, TypeError),
        ("workflow_willingness", 1.0, TypeError),
        ("workflow_willingness", -1, ValueError),
        ("workflow_willingness", 5, ValueError),
    ],
)
def test_ratings_must_be_exact_integers_from_zero_through_four(
    field: str,
    value: object,
    error_type: type[Exception],
) -> None:
    values: dict[str, object] = {
        "process_clarity": 4,
        "camera_smoothness": 3,
        "information_load": 2,
        "workflow_willingness": 1,
    }
    values[field] = value

    with pytest.raises(error_type, match=field):
        build_synthetic_showcase_zip(
            **values,  # type: ignore[arg-type]
            recording_state="saved",
            generated_at=GENERATED_AT,
        )


def test_saved_recording_requires_camera_rating() -> None:
    with pytest.raises(ValueError, match="camera_smoothness"):
        _build_archive(camera_smoothness=None, recording_state="saved")


@pytest.mark.parametrize("recording_state", ["", "recorded", "SAVED", 1, None])
def test_recording_state_is_one_of_three_exact_strings(
    recording_state: object,
) -> None:
    with pytest.raises((TypeError, ValueError), match="recording_state"):
        _build_archive(recording_state=recording_state)  # type: ignore[arg-type]


@pytest.mark.parametrize(
    ("generated_at", "error_type"),
    [
        ("2026-07-28T09:15:30Z", TypeError),
        (datetime(2026, 7, 28, 9, 15, 30), ValueError),
        (
            datetime(
                2026,
                7,
                28,
                18,
                15,
                30,
                tzinfo=timezone(timedelta(hours=9)),
            ),
            ValueError,
        ),
        (
            datetime(
                2026,
                7,
                28,
                9,
                15,
                30,
                1,
                tzinfo=timezone.utc,
            ),
            ValueError,
        ),
    ],
)
def test_generation_time_uses_generic_bundle_utc_contract(
    generated_at: object,
    error_type: type[Exception],
) -> None:
    with pytest.raises(error_type, match="generated_at"):
        _build_archive(generated_at=generated_at)  # type: ignore[arg-type]


def test_generation_time_timezone_failure_does_not_expose_exception_chain() -> None:
    generated_at = datetime(2026, 7, 28, tzinfo=_HostileTimezone())

    with pytest.raises(ValueError, match="generated_at") as error:
        _build_archive(generated_at=generated_at)

    assert "PRIVATE_GENERATION_TIMEZONE_7F3A" not in str(error.value)
    assert error.value.__cause__ is None
    assert error.value.__context__ is None


def test_mutable_timezone_is_replaced_with_a_trusted_plain_utc_copy(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    mutable_timezone = _MutableTimezone()
    generated_at = datetime(
        2026,
        7,
        28,
        9,
        15,
        30,
        tzinfo=mutable_timezone,
    )
    observed: dict[str, object] = {}
    real_builder = showcase_export.build_local_export_bundle

    def inspect_bundle(**kwargs: object) -> object:
        observed["exported_at"] = kwargs["exported_at"]
        return real_builder(**kwargs)  # type: ignore[arg-type]

    monkeypatch.setattr(showcase_export, "build_local_export_bundle", inspect_bundle)

    archive = _build_archive(generated_at=generated_at)
    snapshot, _ = _read_archive(archive)
    trusted_generated_at = observed["exported_at"]

    assert mutable_timezone.utcoffset_calls == 1
    assert trusted_generated_at is not generated_at
    assert type(trusted_generated_at) is datetime
    assert trusted_generated_at.tzinfo is timezone.utc  # type: ignore[union-attr]
    assert trusted_generated_at.utcoffset() == timedelta(0)  # type: ignore[union-attr]
    assert snapshot["generated_at_utc"] == "2026-07-28T09:15:30Z"


def test_datetime_subclass_is_rejected_without_calling_hostile_isoformat() -> None:
    generated_at = _HostileDatetime(
        2026,
        7,
        28,
        9,
        15,
        30,
        tzinfo=timezone.utc,
    )

    with pytest.raises(ValueError, match="generated_at") as error:
        _build_archive(generated_at=generated_at)

    assert "PRIVATE_DATETIME_ABC123" not in str(error.value)
    assert error.value.__cause__ is None
    assert error.value.__context__ is None


def test_timedelta_subclass_comparison_error_is_sanitized_without_a_chain() -> None:
    offset_timezone = _OffsetTimezone(_RaisingOffset())
    generated_at = datetime(2026, 7, 28, tzinfo=offset_timezone)

    with pytest.raises(ValueError, match="generated_at") as error:
        _build_archive(generated_at=generated_at)

    assert "PRIVATE_OFFSET_COMPARE_77" not in str(error.value)
    assert error.value.__cause__ is None
    assert error.value.__context__ is None
    assert offset_timezone.utcoffset_calls == 1


def test_timedelta_subclass_cannot_disguise_a_nonzero_offset() -> None:
    offset_timezone = _OffsetTimezone(_LyingNonzeroOffset())
    generated_at = datetime(2026, 7, 28, tzinfo=offset_timezone)

    with pytest.raises(ValueError, match="generated_at"):
        _build_archive(generated_at=generated_at)

    assert offset_timezone.utcoffset_calls == 1


def test_source_is_closed_to_operational_data_storage_and_network_modules() -> None:
    source = MODULE_SOURCE.read_text(encoding="utf-8")
    tree = ast.parse(source)
    _assert_source_imports_are_isolated(source)
    assert not any(
        isinstance(node, ast.Call)
        and isinstance(node.func, ast.Name)
        and node.func.id in {"open", "exec", "eval", "compile", "__import__"}
        for node in ast.walk(tree)
    )

    lowered = source.casefold()
    forbidden_source_fragments = (
        "participant_id",
        "participantid",
        "media_filename",
        "recording_path",
        "file_path",
        "http://",
        "https://",
    )
    assert not any(fragment in lowered for fragment in forbidden_source_fragments)


@pytest.mark.parametrize(
    "unsafe_module",
    [
        "os",
        "subprocess",
        "httpx",
        "sqlite3",
        "importlib",
        "network",
        "socket",
        "storage",
    ],
)
def test_source_import_allowlist_rejects_every_unapproved_root(
    unsafe_module: str,
) -> None:
    source = MODULE_SOURCE.read_text(encoding="utf-8")

    with pytest.raises(AssertionError):
        _assert_source_imports_are_isolated(f"{source}\nimport {unsafe_module}\n")


@pytest.mark.parametrize(
    ("recording_state", "camera_smoothness"),
    [("saved", 3), ("skipped", None), ("failed", None)],
)
def test_archive_content_contains_only_invented_ids_and_no_private_terms(
    recording_state: str,
    camera_smoothness: int | None,
) -> None:
    archive = _build_archive(
        recording_state=recording_state,
        camera_smoothness=camera_smoothness,
    )
    snapshot, workbook = _read_archive(archive)

    assert tuple(
        item["item_id"] for item in snapshot["ratings"]  # type: ignore[index]
    ) == ITEM_IDS
    assert workbook.sheetnames == ["Session", "Responses", "Recording"]
    _assert_archive_has_no_private_content(archive)


def test_privacy_gate_rejects_private_term_in_raw_json_extra_field() -> None:
    archive = _build_archive()
    json_bytes = _archive_member_bytes(archive, "responses.json")
    assert json_bytes.startswith(b"{")
    mutated_json = (
        b'{"private_marker":"sicq","private_marker":"safe",'
        + json_bytes[1:]
    )
    mutated_archive = _replace_archive_member(
        archive,
        "responses.json",
        mutated_json,
    )

    with pytest.raises(AssertionError, match="sicq"):
        _assert_archive_has_no_private_content(mutated_archive)


def test_privacy_gate_rejects_private_term_in_workbook_metadata() -> None:
    archive = _build_archive()
    workbook = openpyxl.load_workbook(
        BytesIO(_archive_member_bytes(archive, "responses.xlsx"))
    )
    private_term = "\u81ea\u4f24"
    workbook.properties.title = private_term
    workbook_output = BytesIO()
    workbook.save(workbook_output)
    mutated_archive = _replace_archive_member(
        archive,
        "responses.xlsx",
        workbook_output.getvalue(),
    )

    with pytest.raises(AssertionError, match=private_term):
        _assert_archive_has_no_private_content(mutated_archive)


def test_privacy_gate_rejects_private_term_in_non_cell_xml_payload() -> None:
    archive = _build_archive()
    workbook_bytes = _archive_member_bytes(archive, "responses.xlsx")
    mutated_workbook = _add_workbook_xml_payload(
        workbook_bytes,
        "customXml/item1.xml",
        b'<?xml version="1.0" encoding="UTF-8"?><private>fasm</private>',
    )
    mutated_archive = _replace_archive_member(
        archive,
        "responses.xlsx",
        mutated_workbook,
    )

    with pytest.raises(AssertionError, match="fasm"):
        _assert_archive_has_no_private_content(mutated_archive)

import ast
from datetime import date
from io import BytesIO
import json
from pathlib import Path
import re
from zipfile import ZipFile

import openpyxl
import pytest
from streamlit.testing.v1 import AppTest

from questionnaire_specs import (
    DAILY_CONDITIONAL,
    DAILY_CORE,
    FORMAL_INSTRUMENTS,
    VISIT_INSTRUMENT_IDS,
    WEEKLY_INSTRUMENTS,
)
from questionnaire_ui import build_flow, formal_flow, questionnaire_state_keys
from session_record_workflow import (
    create_session_record,
    mark_questionnaire_visit_complete,
    persist_daily_questionnaire,
    persist_formal_questionnaire,
    questionnaire_answers,
    questionnaire_visit_complete,
)


FIXTURE = Path(__file__).parent / "fixtures" / "questionnaire_app.py"
RECORD_DATE = date(2026, 7, 27)
SUBJECT_ID = "sub-001"
TOKEN = "01abcdef"
CREATED_AT_ISO = "2026-07-27T10:00:00+00:00"
COMPLETED_AT_ISO = "2026-07-27T10:15:00+00:00"
EXPORTED_AT_ISO = "2026-07-27T10:30:00+00:00"
SESSION_RECORD_KEY = "session_record"
SESSION_EXPORT_KEY = "session_export"
SESSION_NAMESPACE = "fixture-record"
RECORDING_KEYS = (
    "version",
    "storage",
    "status",
    "mode",
    "duration_seconds",
    "camera_ready",
    "microphone_ready",
    "saved_confirmed",
)
ALLOWED_RUNTIME_CACHE_DIRS = frozenset(
    {".pytest_cache", ".streamlit", "__pycache__"}
)
FORBIDDEN_PARTICIPANT_TERMS = frozenset(
    {
        "derived",
        "risk",
        "risks",
        "score",
        "scored",
        "scores",
        "scoring",
        "threshold",
        "thresholds",
        "path",
        "paths",
        "upload",
        "uploaded",
        "uploading",
        "uploads",
    }
)
PRIVATE_EXPORT_IDENTIFIERS = frozenset(
    {
        "derived_metrics",
        "formal_visits",
        "local_path",
        "raw_answers",
        "record_id",
        "remote_path",
        "revision",
        "safety_signals",
        "video_filename",
    }
)
PARTICIPANT_ARTIFACT_SUFFIXES = frozenset(
    {
        ".avi",
        ".mkv",
        ".mov",
        ".mp4",
        ".mpeg",
        ".mpg",
        ".webm",
        ".xls",
        ".xlsx",
        ".zip",
    }
)
PARTICIPANT_ARTIFACT_NAME_TOKENS = frozenset(
    {
        "answer",
        "answers",
        "participant",
        "questionnaire",
        "recording",
        "recordings",
        "response",
        "responses",
    }
)
PARTICIPANT_DATA_IDENTIFIERS = frozenset(
    {
        "answers",
        "answered_field_ids",
        "conditional_details",
        "daily_core",
        "formal_visits",
        "participant_id",
        "questionnaire_answers",
        "record_id",
        "responses",
        "subject_id",
        "weekly_extension",
    }
)
QUESTIONNAIRE_FIELD_IDS = frozenset(
    question.id
    for question in (
        *DAILY_CORE,
        *DAILY_CONDITIONAL,
        *(
            question
            for instrument in WEEKLY_INSTRUMENTS
            for question in instrument.questions
        ),
        *(
            question
            for instrument in FORMAL_INSTRUMENTS.values()
            for question in instrument.questions
        ),
    )
)
PARTICIPANT_DATA_BYTE_MARKERS = tuple(
    (marker, marker.encode("utf-8"))
    for marker in sorted(
        PARTICIPANT_DATA_IDENTIFIERS | QUESTIONNAIRE_FIELD_IDS
    )
)


def _participant_artifact_reason(path: Path, relative: Path) -> str | None:
    path_tokens = {
        token
        for part in relative.parts
        for token in re.findall(r"[a-z0-9]+", part.casefold())
    }
    if "recordings" in path_tokens:
        return "recordings directory"
    if path.is_dir():
        return None
    if path.suffix.casefold() in PARTICIPANT_ARTIFACT_SUFFIXES:
        return f"prohibited {path.suffix.casefold()} artifact"
    if path.suffix.casefold() in {".pyc", ".pyo"}:
        return None
    filename_tokens = set(
        re.findall(r"[a-z0-9]+", path.name.casefold())
    )
    if (
        "index" in filename_tokens
        and "store" in path_tokens
    ):
        return "store index"
    if path_tokens & PARTICIPANT_ARTIFACT_NAME_TOKENS:
        return "participant artifact name"
    data = path.read_bytes().lower()
    for marker, encoded_marker in PARTICIPANT_DATA_BYTE_MARKERS:
        if encoded_marker in data:
            return f"participant data marker {marker}"
    return None


def _operational_side_effect_snapshot(root: Path) -> dict[str, object]:
    snapshot: dict[str, object] = {}
    for path in sorted(root.rglob("*")):
        relative = path.relative_to(root)
        artifact_reason = _participant_artifact_reason(path, relative)
        assert artifact_reason is None, (
            f"participant artifact detected at {relative.as_posix()}: "
            f"{artifact_reason}"
        )
        if ALLOWED_RUNTIME_CACHE_DIRS.intersection(relative.parts):
            continue
        key = relative.as_posix()
        if path.is_symlink():
            snapshot[key] = ("symlink", str(path.readlink()))
        elif path.is_dir():
            snapshot[key] = ("directory", None)
        else:
            snapshot[key] = ("file", path.read_bytes())
    return snapshot


def _reject_network_calls(monkeypatch) -> None:
    import http.client
    import socket
    import urllib.request

    import requests.sessions

    def reject(*args, **kwargs):
        raise AssertionError("operational fixture attempted network access")

    monkeypatch.setattr(requests.sessions.Session, "request", reject)
    monkeypatch.setattr(urllib.request, "urlopen", reject)
    monkeypatch.setattr(http.client.HTTPConnection, "request", reject)
    monkeypatch.setattr(http.client.HTTPSConnection, "request", reject)
    monkeypatch.setattr(socket, "create_connection", reject)
    monkeypatch.setattr(socket.socket, "connect", reject)
    monkeypatch.setattr(socket.socket, "connect_ex", reject)


def _negative_daily_answers() -> dict[str, object]:
    return {
        "nssi_thought_present_24h": False,
        "nssi_behavior_present_24h": False,
        "suicide_thought_present_24h": False,
        "nssi_urge_now": 0,
        "nssi_resistance_confidence_now": 0,
    }


def _raw_value(question, *, positive: bool) -> object:
    if question.kind == "boolean":
        return positive
    if question.kind == "slider":
        return question.max_value if positive else question.min_value
    if question.kind == "integer":
        return question.min_value
    if question.kind == "multiselect":
        return list(question.options[:2]) if positive else []
    if question.kind == "text":
        return f"raw:{question.id}" if positive else ""
    raise AssertionError(f"unsupported question kind: {question.kind}")


def _positive_daily_answers() -> dict[str, object]:
    questions = (*DAILY_CORE, *DAILY_CONDITIONAL)
    answers = {
        question.id: _raw_value(question, positive=True)
        for question in questions
    }
    first_count = next(
        question.id for question in DAILY_CONDITIONAL if question.kind == "integer"
    )
    answers[first_count] = 1
    return answers


def _weekly_questions():
    return tuple(
        question
        for instrument in WEEKLY_INSTRUMENTS
        for question in instrument.questions
    )


def _formal_questions(visit: str):
    return tuple(
        question
        for instrument_id in VISIT_INSTRUMENT_IDS[visit]
        for question in FORMAL_INSTRUMENTS[instrument_id].questions
    )


def _recording(status: str = "saved") -> dict[str, object]:
    saved = status == "saved"
    return {
        "version": 2,
        "storage": "browser_local",
        "status": status,
        "mode": "long" if saved else "demo",
        "duration_seconds": 60 if saved else 0,
        "camera_ready": saved,
        "microphone_ready": saved,
        "saved_confirmed": saved,
    }


def _new_record(
    *, day: int = 1, visit: str = "daily", recording_status: str = "saved"
) -> dict[str, object]:
    record = create_session_record(
        SUBJECT_ID,
        RECORD_DATE,
        day,
        visit,
        token=TOKEN,
        now_iso=CREATED_AT_ISO,
    )
    record["recording"] = _recording(recording_status)
    return record


def _unique_element(elements, *, key, label):
    matches = [
        element
        for element in elements
        if element.key == key and element.label == label
    ]
    assert len(matches) == 1
    return matches[0]


def _control_for_question(app, question, key):
    collection = {
        "boolean": app.radio,
        "slider": app.slider,
        "integer": app.number_input,
        "multiselect": app.multiselect,
        "text": app.text_area,
    }[question.kind]
    return _unique_element(collection, key=key, label=question.prompt)


def _answer_and_continue(app, question, visit, value):
    keys = questionnaire_state_keys(SESSION_NAMESPACE, visit)
    control = _control_for_question(app, question, keys.widget(question.id))
    if question.kind == "slider" and control.value == value:
        alternate = value + 1 if value < question.max_value else value - 1
        control.set_value(alternate).run()
        control = _control_for_question(app, question, keys.widget(question.id))
    control.set_value(value).run()
    next_button = _unique_element(
        app.button,
        key=keys.next_button,
        label=(
            "检查并提交"
            if not [
                button
                for button in app.button
                if button.key == keys.next_button and button.label == "继续"
            ]
            else "继续"
        ),
    )
    return next_button.click().run()


def _start_fixture(scenario: str = "day1", *, recording_status: str = "saved"):
    app = AppTest.from_file(str(FIXTURE), default_timeout=10)
    app.query_params["scenario"] = scenario
    app.query_params["recording"] = recording_status
    app.run()
    assert not app.exception
    return app


def _complete_fixture(app, questions, visit: str, answers):
    for question in questions:
        app = _answer_and_continue(app, question, visit, answers[question.id])
        assert not app.exception
    return app


def _visible_text(app) -> str:
    values = [str(app.main), str(app.sidebar)]
    for collection_name in (
        "title",
        "header",
        "subheader",
        "caption",
        "markdown",
        "text",
        "info",
        "warning",
        "error",
        "success",
        "json",
        "code",
        "dataframe",
        "table",
        "metric",
        "button",
        "radio",
        "slider",
        "number_input",
        "text_area",
        "multiselect",
        "checkbox",
        "selectbox",
    ):
        for element in getattr(app, collection_name):
            for attribute in ("value", "label", "help", "placeholder"):
                value = getattr(element, attribute, None)
                if value is not None:
                    values.append(str(value))
    return "\n".join(values)


def _worksheet_rows(worksheet) -> list[dict[str, object]]:
    rows = list(worksheet.iter_rows(values_only=True))
    headers = rows[0]
    return [dict(zip(headers, row, strict=True)) for row in rows[1:]]


def _assert_json_cell(actual: object, expected: object) -> None:
    assert isinstance(actual, str)
    assert actual == json.dumps(
        expected,
        ensure_ascii=False,
        allow_nan=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    assert json.loads(actual) == expected


def _participant_bundle(app):
    bundle = app.session_state[SESSION_EXPORT_KEY]
    assert bundle.filename == "session-20260727-103000.zip"
    assert bundle.mime_type == "application/zip"
    with ZipFile(BytesIO(bundle.data)) as archive:
        assert archive.namelist() == ["responses.json", "responses.xlsx"]
        payload = json.loads(archive.read("responses.json"))
        workbook_data = archive.read("responses.xlsx")
    workbook = openpyxl.load_workbook(BytesIO(workbook_data), data_only=False)
    return payload, workbook


def _surface_strings(value):
    if isinstance(value, dict):
        for key, item in value.items():
            yield str(key)
            yield from _surface_strings(item)
    elif isinstance(value, (list, tuple, set, frozenset)):
        for item in value:
            yield from _surface_strings(item)
    elif value is not None:
        yield str(value)


def _privacy_tokens(value: str) -> tuple[str, ...]:
    separated = re.sub(
        r"(?<=[a-z0-9])(?=[A-Z])|(?<=[A-Z])(?=[A-Z][a-z])",
        " ",
        value,
    )
    return tuple(re.findall(r"[a-z0-9]+", separated.casefold()))


def _assert_participant_surfaces_are_raw_only(app, payload, workbook) -> None:
    workbook_values = tuple(
        value
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows(values_only=True)
        for value in row
        if value is not None
    )
    surfaces = {
        "participant-visible text": _visible_text(app),
        "participant JSON": payload,
        "participant XLSX": workbook_values,
    }
    for label, surface in surfaces.items():
        strings = tuple(_surface_strings(surface))
        terms = {
            token
            for value in strings
            for token in _privacy_tokens(value)
        }
        forbidden_terms = terms & FORBIDDEN_PARTICIPANT_TERMS
        assert not forbidden_terms, (
            f"{label} exposes forbidden participant terms: "
            f"{sorted(forbidden_terms)}"
        )
        identifiers = {
            identifier
            for value in strings
            for identifier in re.findall(
                r"[a-z0-9]+(?:_[a-z0-9]+)*", value.casefold()
            )
        }
        private_identifiers = identifiers & PRIVATE_EXPORT_IDENTIFIERS
        assert not private_identifiers, (
            f"{label} exposes private identifiers: "
            f"{sorted(private_identifiers)}"
        )


@pytest.mark.parametrize(
    "sensitive_text",
    (
        "derivedScore",
        "RiskLevel",
        "thresholdValue",
        "filePath",
        "uploadStatus",
        "derived_score",
        "risk-level",
        "threshold value",
        "file_path",
        "upload status",
    ),
)
def test_participant_privacy_tokens_cover_identifier_styles(sensitive_text):
    class VisibleSurface:
        main = sensitive_text
        sidebar = ""

        def __getattr__(self, name):
            return ()

    class EmptyWorkbook:
        worksheets = ()

    with pytest.raises(AssertionError, match="forbidden participant terms"):
        _assert_participant_surfaces_are_raw_only(
            VisibleSurface(),
            {},
            EmptyWorkbook(),
        )


def test_participant_privacy_tokens_do_not_match_innocent_substrings():
    class VisibleSurface:
        main = "scorecard brisk thresholding pathways preupload"
        sidebar = ""

        def __getattr__(self, name):
            return ()

    class EmptyWorkbook:
        worksheets = ()

    _assert_participant_surfaces_are_raw_only(
        VisibleSurface(),
        {},
        EmptyWorkbook(),
    )


def test_visible_text_collector_captures_every_structured_render_channel(tmp_path):
    sentinel = "VISIBLE-CHANNEL-SENTINEL-7F31"
    app_path = tmp_path / "visible_channels.py"
    app_path.write_text(
        "\n".join(
            (
                "import streamlit as st",
                f"sentinel = {sentinel!r}",
                'st.json({"value": sentinel + ":json"})',
                'st.code(sentinel + ":code")',
                'st.dataframe([{"value": sentinel + ":dataframe"}])',
                'st.table([{"value": sentinel + ":table"}])',
                'st.metric(sentinel + ":metric-label", sentinel + ":metric-value")',
                "st.selectbox(",
                '    sentinel + ":selectbox-label",',
                '    (sentinel + ":selectbox-value",),',
                ")",
            )
        ),
        encoding="utf-8",
    )
    app = AppTest.from_file(str(app_path), default_timeout=10).run()

    assert not app.exception
    visible = _visible_text(app)
    for suffix in (
        ":json",
        ":code",
        ":dataframe",
        ":table",
        ":metric-label",
        ":metric-value",
        ":selectbox-label",
        ":selectbox-value",
    ):
        assert f"{sentinel}{suffix}" in visible


def test_questionnaire_fixture_source_is_session_memory_only():
    source = FIXTURE.read_text(encoding="utf-8")
    tree = ast.parse(source)
    imported_roots = {
        alias.name.split(".", 1)[0]
        for node in ast.walk(tree)
        if isinstance(node, ast.Import)
        for alias in node.names
    }
    imported_roots.update(
        node.module.split(".", 1)[0]
        for node in ast.walk(tree)
        if isinstance(node, ast.ImportFrom) and node.module
    )
    imported_names = {
        (node.module, alias.name)
        for node in ast.walk(tree)
        if isinstance(node, ast.ImportFrom) and node.module
        for alias in node.names
    }
    call_names = {
        node.func.id
        if isinstance(node.func, ast.Name)
        else node.func.attr
        for node in ast.walk(tree)
        if isinstance(node, ast.Call)
        and isinstance(node.func, (ast.Name, ast.Attribute))
    }

    required_calls = {
        "build_participant_export",
        "create_session_record",
        "persist_daily_questionnaire",
        "persist_formal_questionnaire",
    }
    assert {
        ("questionnaire_export", "build_participant_export"),
        ("session_record_workflow", "create_session_record"),
        ("session_record_workflow", "persist_daily_questionnaire"),
        ("session_record_workflow", "persist_formal_questionnaire"),
    } <= imported_names
    assert required_calls <= call_names
    assert any(
        isinstance(node, ast.Attribute)
        and isinstance(node.value, ast.Name)
        and node.value.id == "st"
        and node.attr == "session_state"
        for node in ast.walk(tree)
    )
    assert {
        "record_store",
        "questionnaire_fixture_storage",
        "questionnaire_scoring",
        "upload_workflow",
        "pathlib",
        "tempfile",
        "requests",
        "urllib",
        "httpx",
        "socket",
    }.isdisjoint(imported_roots)
    assert {
        "DailyRecordStore",
        "Path",
        "open",
        "mkdtemp",
        "NamedTemporaryFile",
        "TemporaryDirectory",
        "write_text",
        "write_bytes",
        "mkdir",
        "get_or_create",
    }.isdisjoint(call_names)
    forbidden_source = (
        "DailyRecordStore",
        "record_store",
        "questionnaire_fixture_storage",
        "QUESTIONNAIRE_FIXTURE_STORE",
        "video_filename",
        "remote_path",
        "upload",
        "include_derived",
        "derived_metrics",
        "safety_signals",
        "recordings",
        "requests.",
        "urllib.",
        "httpx.",
    )
    assert all(value.casefold() not in source.casefold() for value in forbidden_source)


def test_questionnaire_fixture_never_constructs_a_server_store(tmp_path, monkeypatch):
    import record_store

    def reject_server_store(*args, **kwargs):
        raise AssertionError("fixture attempted server storage")

    monkeypatch.setattr(record_store, "DailyRecordStore", reject_server_store)
    monkeypatch.setenv("QUESTIONNAIRE_FIXTURE_STORE", str(tmp_path))
    app = _start_fixture("day1")

    assert app.session_state[SESSION_RECORD_KEY]["schema_version"] == 5
    assert list(tmp_path.iterdir()) == []


def test_operational_fixture_has_no_filesystem_or_network_side_effects(
    tmp_path, monkeypatch
):
    before = _operational_side_effect_snapshot(tmp_path)
    monkeypatch.chdir(tmp_path)
    _reject_network_calls(monkeypatch)

    answers = _negative_daily_answers()
    flow = build_flow(answers, 1)
    app = _complete_fixture(_start_fixture("day1"), flow, "daily", answers)

    assert SESSION_EXPORT_KEY in app.session_state
    assert _operational_side_effect_snapshot(tmp_path) == before


@pytest.mark.parametrize(
    ("relative_path", "data"),
    (
        pytest.param(
            ".pytest_cache/sub-001.json",
            b'{"participant_id":"sub-001"}',
            id="participant-json",
        ),
        pytest.param(
            ".streamlit/cache.json",
            b'{"value":{"nssi_urge_now":3}}',
            id="questionnaire-field-json",
        ),
        pytest.param(
            ".streamlit/cache.bin",
            b"binary-prefix participant_id binary-suffix",
            id="extension-independent-participant-data",
        ),
        pytest.param(
            ".streamlit/responses.xlsx",
            b"participant workbook",
            id="participant-xlsx",
        ),
        pytest.param(
            "__pycache__/responses.zip",
            b"participant archive",
            id="participant-zip",
        ),
        pytest.param(
            ".pytest_cache/recording.webm",
            b"participant video",
            id="participant-video",
        ),
        pytest.param(
            ".streamlit/recordings/chunk.bin",
            b"participant media",
            id="recordings-directory",
        ),
        pytest.param(
            "__pycache__/record_store/index.json",
            b'{"records":["sub-001"]}',
            id="store-index",
        ),
        pytest.param(
            ".streamlit/record_store/index.db",
            b"database index",
            id="database-store-index",
        ),
        pytest.param(
            ".pytest_cache/cache.csv",
            b"participant_id,answer\nsub-001,true\n",
            id="other-participant-data",
        ),
    ),
)
def test_cache_directories_cannot_hide_participant_artifacts(
    tmp_path, relative_path, data
):
    artifact = tmp_path / relative_path
    artifact.parent.mkdir(parents=True, exist_ok=True)
    artifact.write_bytes(data)

    with pytest.raises(AssertionError, match="participant artifact"):
        _operational_side_effect_snapshot(tmp_path)


def test_cache_scans_complete_file_beyond_one_mebibyte(tmp_path):
    artifact = tmp_path / ".streamlit/late.json"
    artifact.parent.mkdir(parents=True, exist_ok=True)
    artifact.write_bytes(
        b'{"padding":"'
        + b"x" * (1024 * 1024 + 64)
        + b'","participant_id":"sub-001"}'
    )

    with pytest.raises(AssertionError, match="participant artifact"):
        _operational_side_effect_snapshot(tmp_path)


def test_operational_snapshot_allows_only_benign_cache_changes(tmp_path):
    before = _operational_side_effect_snapshot(tmp_path)
    benign_cache_files = {
        ".pytest_cache/v/cache/nodeids": b"[]",
        ".streamlit/config.toml": b"[browser]\ngatherUsageStats = false\n",
        ".streamlit/metrics.csv": b"metric,value\nruns,1\n",
        "__pycache__/questionnaire_app.pyc": b"ordinary bytecode cache",
    }
    for relative_path, data in benign_cache_files.items():
        cache_file = tmp_path / relative_path
        cache_file.parent.mkdir(parents=True, exist_ok=True)
        cache_file.write_bytes(data)

    assert _operational_side_effect_snapshot(tmp_path) == before


@pytest.mark.parametrize(
    "scenario",
    (
        pytest.param("bogus", id="unknown"),
        pytest.param("", id="blank"),
        pytest.param(["day1", "day7"], id="multiple"),
    ),
)
def test_supplied_noncanonical_scenario_fails_closed_without_fixture_state(scenario):
    app = AppTest.from_file(str(FIXTURE), default_timeout=10)
    app.query_params["scenario"] = scenario
    app.run()

    assert not app.exception
    assert SESSION_RECORD_KEY not in app.session_state
    assert SESSION_EXPORT_KEY not in app.session_state
    assert not any(
        key.startswith("fixture_") for key in app.session_state.filtered_state
    )
    visible = _visible_text(app)
    assert "This questionnaire scenario is unavailable." in visible
    for raw_value in scenario if isinstance(scenario, list) else (scenario,):
        if raw_value:
            assert raw_value not in visible
    assert not app.get("download_button")


@pytest.mark.parametrize(
    ("scenario", "visit", "day"),
    (
        ("day1", "daily", 1),
        ("day7", "daily", 7),
        *((visit, visit, 7) for visit in VISIT_INSTRUMENT_IDS),
    ),
)
def test_fresh_browser_session_creates_one_deterministic_raw_v5_record(
    scenario, visit, day
):
    app = _start_fixture(scenario)
    record = app.session_state[SESSION_RECORD_KEY]

    assert record["schema_version"] == 5
    assert record["record_id"] == "sub-001_20260727_01abcdef"
    assert record["subject_id"] == SUBJECT_ID
    assert record["record_date"] == RECORD_DATE.isoformat()
    assert record["intervention_day"] == day
    assert record["visit"] == visit
    assert record["revision"] == 1
    assert record["created_at_iso"] == CREATED_AT_ISO
    assert record["updated_at_iso"] == CREATED_AT_ISO
    assert record["recording"] == _recording()
    assert record["completion"]["status"] == "draft"
    assert questionnaire_answers(record, visit) == {}
    assert {
        "derived_metrics",
        "safety_signals",
        "upload",
        "history",
        "remote_path",
        "local_path",
    }.isdisjoint(record)
    record_keys = [
        str(key)
        for key, value in app.session_state.filtered_state.items()
        if isinstance(value, dict) and value.get("schema_version") == 5
    ]
    assert record_keys == [SESSION_RECORD_KEY]
    assert not any(
        str(key).startswith("fixture_")
        for key in app.session_state.filtered_state
    )


@pytest.mark.parametrize("status", ("saved", "skipped", "failed"))
def test_fixture_uses_only_sanitized_browser_local_v2_metadata(status):
    app = _start_fixture("day1", recording_status=status)
    metadata = app.session_state[SESSION_RECORD_KEY]["recording"]

    assert tuple(metadata) == RECORDING_KEYS
    assert metadata == _recording(status)
    assert {
        "path",
        "filename",
        "bytes",
        "video",
        "blob",
        "upload",
    }.isdisjoint(metadata)


def test_negative_day_one_keeps_false_zero_status_and_completion_metadata():
    record = _new_record()
    stale = {
        **_positive_daily_answers(),
        **_negative_daily_answers(),
    }
    persisted = persist_daily_questionnaire(
        record,
        stale,
        set(stale),
        current_step=len(DAILY_CORE) - 1,
    )

    expected = _negative_daily_answers()
    assert persisted == expected
    assert record["daily_core"] == expected
    assert record["conditional_details"] == {}
    assert record["weekly_extension"] == {}
    assert questionnaire_answers(record, "daily") == expected
    assert record["completion"]["answered_field_ids"]["daily"] == sorted(expected)
    assert all(
        record["field_status"]["daily"][question.id] == "answered"
        for question in DAILY_CORE
    )
    assert all(
        record["field_status"]["daily"][question.id] == "not_applicable"
        for question in DAILY_CONDITIONAL
    )

    mark_questionnaire_visit_complete(
        record,
        "daily",
        completed_at_iso=COMPLETED_AT_ISO,
    )

    assert questionnaire_visit_complete(record, "daily") is True
    assert record["revision"] == 1
    assert record["created_at_iso"] == CREATED_AT_ISO
    assert record["updated_at_iso"] == COMPLETED_AT_ISO
    assert record["completion"]["status"] == "complete"
    assert record["completion"]["questionnaire_visits"]["daily"] == {
        "status": "complete",
        "revision": 1,
        "completed_at_iso": COMPLETED_AT_ISO,
    }
    assert "derived_metrics" not in record
    assert "safety_signals" not in record


def test_positive_daily_browser_branch_preserves_every_nssi_field_then_drops_stale():
    positive = _positive_daily_answers()
    flow = build_flow(positive, 1)
    app = _complete_fixture(_start_fixture("day1"), flow, "daily", positive)
    record = app.session_state[SESSION_RECORD_KEY]

    expected_conditional_ids = {question.id for question in DAILY_CONDITIONAL}
    assert set(record["conditional_details"]) == expected_conditional_ids
    assert questionnaire_answers(record, "daily") == {
        question.id: positive[question.id] for question in flow
    }
    assert all(
        record["field_status"]["daily"][field_id] == "answered"
        for field_id in expected_conditional_ids
    )
    assert record["completion"]["answered_field_ids"]["daily"] == sorted(
        question.id for question in flow
    )

    stale_record = _new_record()
    stale = {**positive, **_negative_daily_answers()}
    persist_daily_questionnaire(
        stale_record,
        stale,
        set(stale),
        current_step=len(DAILY_CORE) - 1,
    )
    assert stale_record["conditional_details"] == {}
    assert all(
        stale_record["field_status"]["daily"][field_id] == "not_applicable"
        for field_id in expected_conditional_ids
    )


def test_day_seven_keeps_daily_and_weekly_raw_inventory_distinct():
    record = _new_record(day=7)
    weekly_questions = _weekly_questions()
    weekly_answers = {
        question.id: _raw_value(question, positive=True)
        for question in weekly_questions
    }
    answers = {**_negative_daily_answers(), **weekly_answers}
    flow = build_flow(answers, 7)

    persist_daily_questionnaire(
        record,
        answers,
        set(answers),
        current_step=len(flow) - 1,
    )

    daily_ids = {question.id for question in DAILY_CORE}
    weekly_ids = {question.id for question in weekly_questions}
    assert set(record["daily_core"]) == daily_ids
    assert set(record["weekly_extension"]) == weekly_ids
    assert not set(record["daily_core"]) & set(record["weekly_extension"])
    assert list(record["weekly_extension"]) == [
        question.id for question in weekly_questions
    ]
    assert all(
        record["field_status"]["daily"][field_id] == "answered"
        for field_id in weekly_ids
    )
    assert questionnaire_answers(record, "daily") == {
        question.id: answers[question.id] for question in flow
    }


def test_day_seven_browser_journey_exports_weekly_raw_answers():
    weekly_questions = _weekly_questions()
    weekly_answers = {
        question.id: _raw_value(question, positive=True)
        for question in weekly_questions
    }
    answers = {**_negative_daily_answers(), **weekly_answers}
    flow = build_flow(answers, 7)

    app = _complete_fixture(
        _start_fixture("day7"),
        flow,
        "daily",
        answers,
    )
    record = app.session_state[SESSION_RECORD_KEY]
    payload, workbook = _participant_bundle(app)

    weekly_ids = [question.id for question in weekly_questions]
    assert questionnaire_visit_complete(record, "daily") is True
    assert record["weekly_extension"] == weekly_answers
    assert payload["intervention_day"] == 7
    assert payload["answered_field_ids"] == [question.id for question in flow]
    weekly_json = [
        item for item in payload["responses"] if item["field_id"] in weekly_ids
    ]
    assert [item["field_id"] for item in weekly_json] == weekly_ids
    assert [item["raw_value"] for item in weekly_json] == [
        weekly_answers[field_id] for field_id in weekly_ids
    ]
    assert {item["instrument_id"] for item in weekly_json} == {
        instrument.id for instrument in WEEKLY_INSTRUMENTS
    }
    weekly_rows = [
        row
        for row in _worksheet_rows(workbook["Responses"])
        if row["field_id"] in weekly_ids
    ]
    assert [row["field_id"] for row in weekly_rows] == weekly_ids
    assert [row["raw_value"] for row in weekly_rows] == [
        weekly_answers[field_id] for field_id in weekly_ids
    ]
    _assert_participant_surfaces_are_raw_only(app, payload, workbook)
    assert len(app.get("download_button")) == 1


@pytest.mark.parametrize("visit", tuple(VISIT_INSTRUMENT_IDS))
def test_every_formal_visit_keeps_complete_instrument_and_item_inventory(visit):
    record = _new_record(day=7, visit=visit)
    questions = _formal_questions(visit)
    answers = {
        question.id: _raw_value(question, positive=True) for question in questions
    }
    active_questions = formal_flow(visit, answers)

    persist_formal_questionnaire(
        record,
        visit,
        answers,
        set(answers),
        current_step=len(active_questions) - 1,
    )

    visit_payload = record["formal_visits"][visit]
    expected_ids = [question.id for question in active_questions]
    assert expected_ids == [question.id for question in questions]
    assert list(visit_payload["raw_answers"]) == expected_ids
    assert visit_payload["raw_answers"] == answers
    assert questionnaire_answers(record, visit) == answers
    for field_id, expected_value in answers.items():
        stored_value = visit_payload["raw_answers"][field_id]
        assert stored_value == expected_value
        assert type(stored_value) is type(expected_value)
    assert tuple(visit_payload["instruments"]) == VISIT_INSTRUMENT_IDS[visit]
    assert visit_payload["complete"] is True
    expected_status = {field_id: "answered" for field_id in expected_ids}
    assert record["field_status"][visit] == expected_status
    assert record["completion"]["answered_field_ids"][visit] == sorted(answers)
    assert record["completion"]["current_step"][visit] == len(expected_ids) - 1
    for instrument_id in VISIT_INSTRUMENT_IDS[visit]:
        payload = visit_payload["instruments"][instrument_id]
        instrument = FORMAL_INSTRUMENTS[instrument_id]
        expected_instrument_ids = [
            question.id for question in instrument.questions
        ]
        expected_instrument_answers = {
            field_id: answers[field_id] for field_id in expected_instrument_ids
        }
        required_count = sum(question.required for question in instrument.questions)
        assert set(payload) == {
            "instrument_id",
            "instrument_version",
            "label",
            "time_window",
            "raw_answers",
            "completeness",
            "complete",
        }
        assert payload["instrument_id"] == instrument_id
        assert payload["instrument_version"] == "1.0"
        assert payload["label"] == instrument.label
        assert payload["time_window"] == instrument.time_window
        assert list(payload["raw_answers"]) == expected_instrument_ids
        assert payload["raw_answers"] == expected_instrument_answers
        for field_id, expected_value in expected_instrument_answers.items():
            stored_value = payload["raw_answers"][field_id]
            assert stored_value == expected_value
            assert type(stored_value) is type(expected_value)
        assert payload["complete"] is True
        assert payload["completeness"] == {
            "answered": required_count,
            "required": required_count,
        }

    mark_questionnaire_visit_complete(
        record,
        visit,
        completed_at_iso=COMPLETED_AT_ISO,
    )
    assert questionnaire_visit_complete(record, visit) is True
    assert record["completion"]["status"] == "complete"
    assert record["completion"]["answered_field_ids"][visit] == sorted(answers)
    assert record["field_status"][visit] == expected_status
    assert record["completion"]["questionnaire_visits"][visit] == {
        "status": "complete",
        "revision": 1,
        "completed_at_iso": COMPLETED_AT_ISO,
    }


@pytest.mark.parametrize("visit", tuple(VISIT_INSTRUMENT_IDS))
def test_every_formal_visit_completes_browser_journey_to_local_export(visit):
    questions = _formal_questions(visit)
    answers = {
        question.id: _raw_value(question, positive=True) for question in questions
    }
    flow = formal_flow(visit, answers)
    assert [question.id for question in flow] == [
        question.id for question in questions
    ]

    app = _complete_fixture(_start_fixture(visit), flow, visit, answers)
    record = app.session_state[SESSION_RECORD_KEY]
    bundle = app.session_state[SESSION_EXPORT_KEY]

    assert record["visit"] == visit
    assert questionnaire_visit_complete(record, visit) is True
    assert record["completion"]["questionnaire_visits"][visit] == {
        "status": "complete",
        "revision": 1,
        "completed_at_iso": COMPLETED_AT_ISO,
    }
    assert questionnaire_answers(record, visit) == answers
    assert len(app.get("download_button")) == 1
    assert bundle.filename == "session-20260727-103000.zip"
    assert bundle.mime_type == "application/zip"
    with ZipFile(BytesIO(bundle.data)) as archive:
        assert archive.namelist() == ["responses.json", "responses.xlsx"]
        payload = json.loads(archive.read("responses.json"))
        workbook_data = archive.read("responses.xlsx")

    assert payload["visit"] == visit
    assert payload["answered_field_ids"] == [question.id for question in questions]
    assert [item["instrument_id"] for item in payload["visits"]] == list(
        VISIT_INSTRUMENT_IDS[visit]
    )
    exported_answers = {
        item["field_id"]: item["raw_value"] for item in payload["responses"]
    }
    assert exported_answers == answers
    workbook = openpyxl.load_workbook(BytesIO(workbook_data), data_only=False)
    _assert_participant_surfaces_are_raw_only(app, payload, workbook)


def test_browser_required_gate_and_back_navigation_restore_answer_and_step():
    app = _start_fixture("day1")
    keys = questionnaire_state_keys(SESSION_NAMESPACE, "daily")
    _unique_element(app.button, key=keys.next_button, label="继续").click().run()

    assert app.error
    assert app.session_state[keys.step] == 0
    assert questionnaire_answers(app.session_state[SESSION_RECORD_KEY], "daily") == {}

    app = _answer_and_continue(app, DAILY_CORE[0], "daily", False)
    assert app.session_state[keys.step] == 1
    _unique_element(app.button, key=keys.back_button, label="←").click().run()

    assert app.session_state[keys.step] == 0
    assert app.session_state[SESSION_RECORD_KEY]["completion"]["current_step"][
        "daily"
    ] == 0
    assert questionnaire_answers(
        app.session_state[SESSION_RECORD_KEY], "daily"
    ) == {DAILY_CORE[0].id: False}
    restored = _control_for_question(app, DAILY_CORE[0], keys.widget(DAILY_CORE[0].id))
    assert restored.value is False


def test_browser_fixture_retains_same_session_but_refresh_before_download_loses_draft():
    app = _start_fixture("day1")
    app = _answer_and_continue(app, DAILY_CORE[0], "daily", False)
    retained_record = app.session_state[SESSION_RECORD_KEY]
    stable_id = retained_record["record_id"]

    app.run()
    keys = questionnaire_state_keys(SESSION_NAMESPACE, "daily")
    assert app.session_state[SESSION_RECORD_KEY]["record_id"] == stable_id
    assert questionnaire_answers(
        app.session_state[SESSION_RECORD_KEY], "daily"
    ) == {DAILY_CORE[0].id: False}
    assert app.session_state[keys.step] == 1
    assert SESSION_EXPORT_KEY not in app.session_state

    fresh = _start_fixture("day1")
    fresh_keys = questionnaire_state_keys(SESSION_NAMESPACE, "daily")
    assert fresh.session_state[SESSION_RECORD_KEY]["record_id"] == stable_id
    assert questionnaire_answers(
        fresh.session_state[SESSION_RECORD_KEY], "daily"
    ) == {}
    assert fresh.session_state[fresh_keys.step] == 0
    assert SESSION_EXPORT_KEY not in fresh.session_state


def test_fixture_shows_support_copy_for_answered_suicide_thought_branch():
    app = _start_fixture("day1")
    for question, value in zip(DAILY_CORE[:3], (False, False, True), strict=True):
        app = _answer_and_continue(app, question, "daily", value)
        assert not app.exception

    record = app.session_state[SESSION_RECORD_KEY]
    assert questionnaire_answers(record, "daily")[
        "suicide_thought_present_24h"
    ] is True
    assert "safety_signals" not in record
    visible = _visible_text(app).casefold()
    assert "study support team" in visible
    assert "local emergency services" in visible
    assert "risk" not in visible


def test_current_visit_export_has_exact_json_xlsx_raw_inventory_and_private_surface():
    answers = _positive_daily_answers()
    flow = build_flow(answers, 1)
    app = _complete_fixture(
        _start_fixture("day1"),
        flow,
        "daily",
        answers,
    )
    record = app.session_state[SESSION_RECORD_KEY]
    bundle = app.session_state[SESSION_EXPORT_KEY]

    assert questionnaire_visit_complete(record, "daily") is True
    assert record["completion"]["questionnaire_visits"]["daily"] == {
        "status": "complete",
        "revision": 1,
        "completed_at_iso": COMPLETED_AT_ISO,
    }
    assert bundle.filename == "session-20260727-103000.zip"
    assert bundle.mime_type == "application/zip"
    with ZipFile(BytesIO(bundle.data)) as archive:
        assert archive.namelist() == ["responses.json", "responses.xlsx"]
        payload = json.loads(archive.read("responses.json"))
        workbook_data = archive.read("responses.xlsx")
    workbook = openpyxl.load_workbook(BytesIO(workbook_data), data_only=False)

    export_questions = (*DAILY_CORE, *DAILY_CONDITIONAL)
    expected_answered = [question.id for question in export_questions]
    assert payload["export_schema_version"] == 1
    assert payload["participant_id"] == SUBJECT_ID
    assert payload["record_date"] == RECORD_DATE.isoformat()
    assert payload["intervention_day"] == 1
    assert payload["visit"] == "daily"
    assert payload["exported_at_iso"] == EXPORTED_AT_ISO
    assert payload["recording"] == _recording()
    assert payload["answered_field_ids"] == expected_answered
    assert payload["field_status"] == record["field_status"]["daily"]
    assert [item["instrument_id"] for item in payload["visits"]] == [
        "daily_nssi_ema"
    ]
    assert payload["visits"][0]["visit_status"] == "complete"
    assert payload["visits"][0]["instrument_status"] == "complete"
    assert payload["visits"][0]["completed_at_iso"] == COMPLETED_AT_ISO
    json_responses = {
        (item["visit"], item["instrument_id"], item["field_id"]): item
        for item in payload["responses"]
    }
    expected_response_keys = {
        ("daily", "daily_nssi_ema", question.id) for question in export_questions
    }
    assert set(json_responses) == expected_response_keys
    for question in export_questions:
        key = ("daily", "daily_nssi_ema", question.id)
        item = json_responses[key]
        expected_value = answers[question.id]
        assert item["instrument_version"] == "1.0"
        assert item["question_text"] == question.prompt
        assert item["question_kind"] == question.kind
        assert item["answered"] is True
        assert item["applicability"] == "applicable"
        assert item["raw_value"] == expected_value
        assert type(item["raw_value"]) is type(expected_value)
        assert isinstance(item["display_value"], str)

    assert workbook.sheetnames == ["Session", "Responses", "Visits", "Recording"]
    session_rows = _worksheet_rows(workbook["Session"])
    response_rows = _worksheet_rows(workbook["Responses"])
    visit_rows = _worksheet_rows(workbook["Visits"])
    recording_rows = _worksheet_rows(workbook["Recording"])
    assert len(session_rows) == 1
    session_row = session_rows[0]
    assert set(session_row) == {
        "export_schema_version",
        "participant_id",
        "record_date",
        "intervention_day",
        "visit",
        "exported_at_iso",
        "daily_context",
        "answered_field_ids",
        "field_status",
    }
    assert session_row["export_schema_version"] == 1
    assert session_row["participant_id"] == SUBJECT_ID
    assert session_row["record_date"] == RECORD_DATE.isoformat()
    assert session_row["intervention_day"] == 1
    assert session_row["visit"] == "daily"
    assert session_row["exported_at_iso"] == EXPORTED_AT_ISO
    _assert_json_cell(session_row["daily_context"], payload["daily_context"])
    _assert_json_cell(session_row["answered_field_ids"], expected_answered)
    _assert_json_cell(session_row["field_status"], payload["field_status"])

    workbook_responses = {
        (row["visit"], row["instrument_id"], row["field_id"]): row
        for row in response_rows
    }
    assert set(workbook_responses) == expected_response_keys
    for key, json_item in json_responses.items():
        row = workbook_responses[key]
        assert set(row) == {
            "visit",
            "instrument_id",
            "field_id",
            "question_text",
            "question_kind",
            "answered",
            "applicability",
            "raw_value",
            "display_value",
        }
        for field_name in (
            "visit",
            "instrument_id",
            "field_id",
            "question_text",
            "question_kind",
            "answered",
            "applicability",
            "display_value",
        ):
            assert row[field_name] == json_item[field_name]
        expected_value = json_item["raw_value"]
        if isinstance(expected_value, (list, dict)):
            _assert_json_cell(row["raw_value"], expected_value)
        else:
            assert row["raw_value"] == expected_value
            assert type(row["raw_value"]) is type(expected_value)

    json_visits = {
        (item["visit"], item["instrument_id"]): item
        for item in payload["visits"]
    }
    workbook_visits = {
        (row["visit"], row["instrument_id"]): row for row in visit_rows
    }
    assert set(json_visits) == {("daily", "daily_nssi_ema")}
    assert set(workbook_visits) == set(json_visits)
    for key, json_item in json_visits.items():
        row = workbook_visits[key]
        assert set(row) == {
            "visit",
            "visit_status",
            "completed_at_iso",
            "instrument_id",
            "instrument_version",
            "instrument_status",
            "answered_field_ids",
            "field_status",
        }
        for field_name in (
            "visit",
            "visit_status",
            "completed_at_iso",
            "instrument_id",
            "instrument_version",
            "instrument_status",
        ):
            assert row[field_name] == json_item[field_name]
        _assert_json_cell(
            row["answered_field_ids"], json_item["answered_field_ids"]
        )
        _assert_json_cell(row["field_status"], json_item["field_status"])

    assert recording_rows == [_recording()]
    assert tuple(recording_rows[0]) == RECORDING_KEYS
    assert recording_rows[0]["status"] == "saved"

    _assert_participant_surfaces_are_raw_only(app, payload, workbook)

    assert len(app.get("download_button")) == 1

    finish = _unique_element(
        app.button,
        key="session_finish",
        label="Finish this session",
    )
    assert finish.disabled is True
    app.session_state["session_recorder::pending"] = _recording("failed")
    app.session_state["session_context::narrative"] = "sensitive draft"
    _unique_element(
        app.checkbox,
        key="session_saved_locally",
        label="I confirm the questionnaire ZIP is saved locally",
    ).check().run()
    finish = _unique_element(
        app.button,
        key="session_finish",
        label="Finish this session",
    )
    assert finish.disabled is False
    finish.click().run()

    remaining = {
        str(key)
        for key in app.session_state.filtered_state
        if not str(key).startswith("$$")
    }
    assert remaining == {"session_complete"}
    assert app.session_state["session_complete"] is True
    assert "This session is complete." in _visible_text(app)
    assert not app.get("download_button")
    assert not app.radio
    assert not app.slider
